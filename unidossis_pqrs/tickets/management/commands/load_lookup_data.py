import os
import openpyxl
import json
import urllib.request
from django.core.management.base import BaseCommand
from tickets.models import Cliente, Ciudad, Cargo

class Command(BaseCommand):
    help = 'Carga datos de clientes, ciudades y cargos desde archivos y listas predefinidas'

    def handle(self, *args, **options):
        # 1. Cargar Clientes desde Excel
        excel_path = os.path.join(os.getcwd(), 'Listado de Cliente_Nombre comercial.xlsx')
        if os.path.exists(excel_path):
            self.stdout.write(self.style.SUCCESS(f'Leyendo clientes desde {excel_path}...'))
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            clientes_agregados = 0
            # Empezamos en la fila 2 para saltar el encabezado 'Cliente'
            for row in range(2, sheet.max_row + 1):
                nombre_cliente = sheet.cell(row=row, column=1).value
                if nombre_cliente:
                    nombre_cliente = str(nombre_cliente).strip()
                    # Evitar duplicados
                    obj, created = Cliente.objects.get_or_create(nombre=nombre_cliente)
                    if created:
                        clientes_agregados += 1
            self.stdout.write(self.style.SUCCESS(f'Se agregaron {clientes_agregados} nuevos clientes.'))
        else:
            self.stdout.write(self.style.WARNING(f'No se encontró el archivo {excel_path}'))

        # 2. Cargar Ciudades de Colombia (Base Completa >1100 municipios)
        json_url = "https://raw.githubusercontent.com/marcovega/colombia-json/master/colombia.json"
        try:
            self.stdout.write(self.style.SUCCESS(f'Descargando base de datos completa de municipios desde {json_url}...'))
            with urllib.request.urlopen(json_url) as response:
                if response.status == 200:
                    data = json.loads(response.read().decode())
                    ciudades_agregadas = 0
                    for depto in data:
                        nombre_depto = depto.get('departamento', '')
                        for ciudad_nom in depto.get('ciudades', []):
                            # Estructura: "Palmira (Valle del Cauca)" o similar
                            nombre_completo = f"{ciudad_nom} ({nombre_depto})"
                            obj, created = Ciudad.objects.get_or_create(nombre=nombre_completo)
                            if created:
                                ciudades_agregadas += 1
                    self.stdout.write(self.style.SUCCESS(f'Se agregaron {ciudades_agregadas} nuevos municipios/ciudades.'))
                else:
                    self.stdout.write(self.style.ERROR('No se pudo descargar el archivo JSON.'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error al cargar municipios: {str(e)}'))

        # 3. Cargar Cargos
        cargos = [
            "Director Técnico", "Director de Calidad", "Químico Farmacéutico", 
            "Regente", "Enfermera", "Médico"
        ]
        cargos_agregados = 0
        for cargo_nom in cargos:
            obj, created = Cargo.objects.get_or_create(nombre=cargo_nom)
            if created:
                cargos_agregados += 1
        self.stdout.write(self.style.SUCCESS(f'Se agregaron {cargos_agregados} nuevos cargos.'))
