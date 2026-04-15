"""
Management command to analyze the PQRS Excel file structure.
Usage: python manage.py importar_pqrs --analyze
       python manage.py importar_pqrs --import
"""
import openpyxl
from django.core.management.base import BaseCommand
from django.utils import timezone
from tickets.models import Ticket, Cliente
import os
from datetime import datetime


EXCEL_PATH = r'c:\Users\Francisco Alvarez\App_PQRS_Unidossis\AC-FR-056 LISTADO MAESTRO DE PQRS 2026.xlsx'


class Command(BaseCommand):
    help = 'Analiza e importa PQRS desde el Listado Maestro Excel'

    def add_arguments(self, parser):
        parser.add_argument('--analyze', action='store_true', help='Solo analizar estructura')
        parser.add_argument('--import', action='store_true', dest='do_import', help='Importar tickets')

    def handle(self, *args, **options):
        if not os.path.exists(EXCEL_PATH):
            self.stderr.write(f'No se encontro el archivo: {EXCEL_PATH}')
            return

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        
        self.stdout.write(f'Hojas: {wb.sheetnames}')
        
        # Buscar hoja LISTADO MAESTRO
        sheet = None
        for name in wb.sheetnames:
            if 'LISTADO' in name.upper():
                sheet = wb[name]
                break
        if not sheet:
            sheet = wb[wb.sheetnames[0]]
        
        self.stdout.write(f'Hoja seleccionada: {sheet.title}')
        self.stdout.write(f'Filas: {sheet.max_row}, Columnas: {sheet.max_column}')
        
        # Mostrar encabezados
        self.stdout.write('\n--- ENCABEZADOS ---')
        for row_idx in range(1, min(8, sheet.max_row + 1)):
            vals = []
            for col in range(1, min(sheet.max_column + 1, 25)):
                val = sheet.cell(row=row_idx, column=col).value
                if val is not None:
                    vals.append(f'C{col}={str(val)[:40]}')
            if vals:
                self.stdout.write(f'Fila {row_idx}: {" | ".join(vals)}')

        # Mostrar datos de muestra
        self.stdout.write('\n--- PRIMERAS FILAS DE DATOS ---')
        # Find header row
        header_row = 1
        for r in range(1, 10):
            for c in range(1, 15):
                v = sheet.cell(row=r, column=c).value
                if v and str(v).strip().upper() in ['CODIGO', 'COD', 'FECHA DE RECEPCION', 'FECHA', 'NO']:
                    header_row = r
                    break
            if header_row > 1:
                break

        self.stdout.write(f'Header row: {header_row}')
        
        # Print first 5 data rows
        for data_row in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
            self.stdout.write(f'\n-- Fila {data_row} --')
            for col in range(1, min(sheet.max_column + 1, 25)):
                val = sheet.cell(row=data_row, column=col).value
                hdr = sheet.cell(row=header_row, column=col).value
                if val is not None and hdr is not None:
                    self.stdout.write(f'  {hdr}: {val}')

        if not options.get('do_import'):
            self.stdout.write('\nUse --import para importar los datos')
            return
