from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.db.models import Q, Count, Avg
from django.utils import timezone
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from functools import wraps
import json

from ..models import (
    Ticket, ArchivoAdjunto, Cliente, Ciudad, Cargo, MaestroInstitucion,
    PerfilUsuario, ConfiguracionSLA, AlertaSLA, LogActividad,
    ComentarioTicket, EncuestaSatisfaccion, FeedbackIA, IntentoLogin,
    SolicitudResetPassword
)
from ..ia_engine import analizar_ticket_con_ia, conversar_con_analista_ia, reclasificar_ticket_con_ia, generar_resumen_cliente
from ._helpers import rol_requerido, _get_client_ip, _enviar_acuse_recibo, _enviar_respuesta_formal_y_csat

@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def configurar_sla_view(request):
    """Panel para parametrizar el SLA. Los destinatarios se obtienen de los usuarios internos."""
    config = ConfiguracionSLA.objects.filter(activo=True).first()
    mensaje = None

    if request.method == 'POST':
        dias_peligro = int(request.POST.get('dias_alerta_peligro', 11))
        dias_vencido = int(request.POST.get('dias_alerta_vencido', 15))

        if config:
            config.dias_alerta_peligro = dias_peligro
            config.dias_alerta_vencido = dias_vencido
            config.save()
        else:
            config = ConfiguracionSLA.objects.create(
                dias_alerta_peligro=dias_peligro,
                dias_alerta_vencido=dias_vencido,
                activo=True
            )
        mensaje = {'tipo': 'exito', 'texto': 'Configuración SLA guardada exitosamente.'}

    # Destinatarios automáticos desde Usuarios Internos
    directores_regionales = PerfilUsuario.objects.filter(
        rol='director_regional'
    ).select_related('user').order_by('regional')

    admins_pqrs = PerfilUsuario.objects.filter(
        rol__in=['admin_pqrs', 'superadmin']
    ).select_related('user').order_by('user__first_name')

    return render(request, 'tickets/configurar_sla.html', {
        'config': config,
        'mensaje': mensaje,
        'perfil': request.user.perfil,
        'nav_active': 'sla',
        'directores_regionales': directores_regionales,
        'admins_pqrs': admins_pqrs,
    })


@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional')
def reportes_view(request):
    """Dashboard de analytics avanzado con KPIs interactivos."""
    perfil = request.user.perfil

    # Base query según rol
    if perfil.rol in ['superadmin', 'admin_pqrs']:
        qs = Ticket.objects.all()
    else:
        qs = Ticket.objects.filter(regional=perfil.regional)

    # ─ Filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    if fecha_inicio:
        qs = qs.filter(fecha_ingreso__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_ingreso__date__lte=fecha_fin)

    total = qs.count()
    resueltos = qs.filter(estado='resuelto').count()
    abiertos = qs.filter(estado='abierto').count()
    en_revision = qs.filter(estado='revision').count()
    cancelados = qs.filter(estado='cancelado').count()
    cumplimiento_sla = round((resueltos / total * 100), 1) if total > 0 else 0

    # Calcular vencidos sobre abiertos reales
    tickets_abiertos_list = list(qs.exclude(estado__in=['resuelto', 'cancelado']))
    vencidos_count = sum(1 for t in tickets_abiertos_list if t.estado_sla() == 'vencido')
    peligro_count = sum(1 for t in tickets_abiertos_list if t.estado_sla() == 'peligro')

    # ─ Por tipificación (top 8)
    por_tipificacion = list(
        qs.exclude(tipificacion__isnull=True).values('tipificacion')
        .annotate(total=Count('id')).order_by('-total')[:8]
    )

    # ─ Por regional
    por_regional = list(
        qs.exclude(regional__isnull=True).values('regional')
        .annotate(total=Count('id')).order_by('-total')
    )
    # Mapear códigos a labels
    regional_map = dict(Ticket.REGIONAL_CHOICES)
    for r in por_regional:
        r['label'] = regional_map.get(r['regional'], r['regional'])

    # ─ Por tipo de solicitud
    por_tipo = list(
        qs.values('tipo_solicitud').annotate(total=Count('id')).order_by('-total')
    )
    tipo_map = dict(Ticket.TIPO_SOLICITUD_CHOICES)
    for t in por_tipo:
        t['label'] = tipo_map.get(t['tipo_solicitud'], t['tipo_solicitud'])

    # ─ Por criticidad
    por_criticidad = list(
        qs.exclude(criticidad__isnull=True).values('criticidad')
        .annotate(total=Count('id')).order_by('-total')
    )
    criticidad_map = dict(Ticket.CRITICIDAD_CHOICES)
    for c in por_criticidad:
        c['label'] = criticidad_map.get(c['criticidad'], c['criticidad'])

    # ─ Tendencia mensual (últimos 6 meses)
    from django.db.models.functions import TruncMonth
    tendencia = list(
        qs.annotate(mes=TruncMonth('fecha_ingreso'))
        .values('mes').annotate(total=Count('id')).order_by('mes')
    )

    # ─ CSAT promedio
    encuestas = EncuestaSatisfaccion.objects.filter(
        ticket__in=qs, estado='satisfecho', puntuacion__isnull=False
    )
    csat_promedio = encuestas.aggregate(avg=Avg('puntuacion'))['avg']
    csat_promedio = round(csat_promedio, 1) if csat_promedio else None
    total_encuestas = encuestas.count()

    context = {
        'perfil': perfil,
        'total': total,
        'resueltos': resueltos,
        'abiertos': abiertos,
        'en_revision': en_revision,
        'cancelados': cancelados,
        'cumplimiento_sla': cumplimiento_sla,
        'vencidos_count': vencidos_count,
        'peligro_count': peligro_count,
        'csat_promedio': csat_promedio,
        'total_encuestas': total_encuestas,
        # Datos JSON para Chart.js
        'por_tipificacion_json': json.dumps(por_tipificacion, default=str),
        'por_regional_json': json.dumps(por_regional, default=str),
        'por_tipo_json': json.dumps(por_tipo, default=str),
        'por_criticidad_json': json.dumps(por_criticidad, default=str),
        'tendencia_json': json.dumps(tendencia, default=str),
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'nav_active': 'reportes',
    }
    return render(request, 'tickets/reportes.html', context)


@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional')
def exportar_excel_view(request):
    """Exporta los tickets a Excel con formato corporativo Unidossis."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecute: pip install openpyxl', status=500)

    perfil = request.user.perfil
    if perfil.rol in ['superadmin', 'admin_pqrs']:
        qs = Ticket.objects.all()
    else:
        qs = Ticket.objects.filter(regional=perfil.regional)

    # Filtros opcionales
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    regional_filtro = request.GET.get('regional')
    if fecha_inicio:
        qs = qs.filter(fecha_ingreso__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_ingreso__date__lte=fecha_fin)
    if regional_filtro:
        qs = qs.filter(regional=regional_filtro)

    qs = qs.order_by('-fecha_ingreso')

    # ─ Crear libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte PQRS UNIDOSSIS'

    # ─ Colores corporativos
    AZUL_CORP = 'FF0F3460'
    ROJO_CORP = 'FFe63946'
    GRIS_HEADER = 'FF1E293B'
    GRIS_CLARO = 'FFF8FAFC'
    BLANCO = 'FFFFFFFF'

    # ─ Fila 1: Título corporativo
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = 'REPORTE DE PQRS — UNIDOSSIS S.A.S.'
    title_cell.font = Font(name='Calibri', size=14, bold=True, color=BLANCO)
    title_cell.fill = PatternFill('solid', fgColor=AZUL_CORP)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ─ Fila 2: Fecha del reporte
    ws.merge_cells('A2:O2')
    date_cell = ws['A2']
    date_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} | Por: {request.user.get_full_name() or request.user.username}'
    date_cell.font = Font(name='Calibri', size=10, color='FF64748B')
    date_cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # ─ Fila 3: Encabezados
    headers = [
        ('ID Caso', 12), ('Fecha Ingreso', 16), ('Tipo Solicitud', 16),
        ('Estado', 13), ('Criticidad', 14), ('Cliente / Institución', 28),
        ('Ciudad', 16), ('Solicitante', 22), ('Correo', 28),
        ('Regional', 18), ('Área / Proceso', 18), ('Línea de Servicio', 18),
        ('Tipificación', 24), ('Responsable', 20), ('Días Transcurridos', 10),
    ]

    thin = Side(style='thin', color='FFE2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header_text
        cell.font = Font(name='Calibri', size=10, bold=True, color=BLANCO)
        cell.fill = PatternFill('solid', fgColor=GRIS_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[3].height = 35

    # ─ Mapas de labels
    tipo_map = dict(Ticket.TIPO_SOLICITUD_CHOICES)
    estado_map = dict(Ticket.STATUS_CHOICES)
    criticidad_map = dict(Ticket.CRITICIDAD_CHOICES)
    regional_map = dict(Ticket.REGIONAL_CHOICES)
    area_map = dict(Ticket.AREA_CHOICES)
    linea_map = dict(Ticket.LINEA_CHOICES)
    tipificacion_map = dict(Ticket.TIPIFICACION_CHOICES)

    # ─ Colores de estado
    COLOR_ESTADO = {
        'abierto': 'FF3B82F6',
        'revision': 'FFF59E0B',
        'resuelto': 'FF10B981',
        'cancelado': 'FF94A3B8',
    }
    COLOR_CRITICIDAD = {
        'critica': 'FFEF4444',
        'mayor': 'FFF97316',
        'menor': 'FFF59E0B',
        'informativa': 'FF10B981',
    }

    # ─ Datos
    for row_idx, ticket in enumerate(qs, start=4):
        fill_row = PatternFill('solid', fgColor=BLANCO if row_idx % 2 == 0 else 'FFF8FAFC')
        row_data = [
            ticket.ticket_id,
            ticket.fecha_ingreso.strftime('%d/%m/%Y %H:%M'),
            tipo_map.get(ticket.tipo_solicitud, ticket.tipo_solicitud or ''),
            estado_map.get(ticket.estado, ticket.estado or ''),
            criticidad_map.get(ticket.criticidad, ticket.criticidad or ''),
            ticket.entidad_cliente or ticket.institucion or '',
            ticket.ciudad or '',
            ticket.remitente_nombre or '',
            ticket.remitente_email or '',
            regional_map.get(ticket.regional, ticket.regional or ''),
            area_map.get(ticket.proceso, ticket.proceso or ''),
            linea_map.get(ticket.linea_servicio, ticket.linea_servicio or ''),
            tipificacion_map.get(ticket.tipificacion, ticket.tipificacion or ''),
            ticket.responsable or '',
            ticket.dias_transcurridos(),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=9)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.fill = fill_row

            # Color especial para estado
            if col_idx == 4 and ticket.estado in COLOR_ESTADO:
                cell.font = Font(name='Calibri', size=9, bold=True, color=BLANCO)
                cell.fill = PatternFill('solid', fgColor=COLOR_ESTADO[ticket.estado])
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Color especial para criticidad
            if col_idx == 5 and ticket.criticidad in COLOR_CRITICIDAD:
                cell.font = Font(name='Calibri', size=9, bold=True, color=BLANCO)
                cell.fill = PatternFill('solid', fgColor=COLOR_CRITICIDAD[ticket.criticidad])
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A4'

    # ─ Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'PQRS_UNIDOSSIS_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def encuesta_csat_view(request, token):
    """Vista pública para que el cliente responda la encuesta de satisfacción."""
    encuesta = get_object_or_404(EncuestaSatisfaccion, token=token)

    if encuesta.estado in ['satisfecho', 'insatisfecho']:
        return render(request, 'tickets/encuesta_ya_respondida.html', {'encuesta': encuesta})

    if request.method == 'POST':
        puntuacion = int(request.POST.get('puntuacion', 0))
        comentario = request.POST.get('comentario', '').strip()

        encuesta.puntuacion = puntuacion
        encuesta.comentario_cliente = comentario
        encuesta.fecha_respuesta = timezone.now()

        ticket = encuesta.ticket

        if puntuacion >= 3:
            # Satisfecho → cierre definitivo
            encuesta.estado = 'satisfecho'
            ticket.estado = 'resuelto'
            ticket.save()
            LogActividad.objects.create(
                ticket=ticket, usuario=None,
                accion=f'Encuesta CSAT respondida: SATISFECHO ({puntuacion}/5 ⭐)',
                detalle=f'Comentario: {comentario[:100]}' if comentario else 'Sin comentario'
            )
        else:
            # Insatisfecho → reabrir y alertar
            encuesta.estado = 'insatisfecho'
            ticket.estado = 'revision'
            ticket.save()
            LogActividad.objects.create(
                ticket=ticket, usuario=None,
                accion=f'🚨 Encuesta CSAT: INSATISFECHO ({puntuacion}/5 ⭐) — Ticket REABIERTO automáticamente',
                detalle=f'Comentario cliente: {comentario[:200]}' if comentario else 'Sin comentario'
            )
            # Notificar al equipo que el cliente está insatisfecho
            try:
                config = ConfiguracionSLA.objects.filter(activo=True).first()
                destinatarios = config.get_emails_vencido() if config else []
                if destinatarios:
                    send_mail(
                        subject=f'🚨 CSAT NEGATIVO — Ticket {ticket.ticket_id} reabierto automáticamente',
                        message=(
                            f'El cliente {ticket.remitente_nombre or ticket.remitente_email} '
                            f'respondió la encuesta CSAT con una puntuación de {puntuacion}/5.\n\n'
                            f'El ticket #{ticket.ticket_id} ha sido REABIERTO automáticamente y '
                            f'requiere revisión urgente.\n\n'
                            f'Comentario del cliente: {comentario or "(Sin comentario)"}\n\n'
                            f'— UNIDOSSIS PQRS · Sistema Automático'
                        ),
                        from_email='UNIDOSSIS PQRS <alertas@unidossis.com.co>',
                        recipient_list=destinatarios,
                        fail_silently=True,
                    )
            except Exception:
                pass

        encuesta.save()
        return render(request, 'tickets/encuesta_gracias.html', {
            'encuesta': encuesta,
            'satisfecho': encuesta.estado == 'satisfecho'
        })

    return render(request, 'tickets/encuesta_csat.html', {'encuesta': encuesta})


@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def encuesta_csat_preview(request):
    """Vista de preview del formulario de encuesta CSAT — sin datos reales."""
    return render(request, 'tickets/encuesta_csat.html', {
        'encuesta': None,
        'preview_mode': True,
    })
