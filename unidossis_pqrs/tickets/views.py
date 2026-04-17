from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.db.models import Q, Count, Avg
from django.utils import timezone
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from functools import wraps
import json

from .models import (
    Ticket, ArchivoAdjunto, Cliente, Ciudad, Cargo, MaestroInstitucion,
    PerfilUsuario, ConfiguracionSLA, AlertaSLA, LogActividad,
    ComentarioTicket, EncuestaSatisfaccion, FeedbackIA, IntentoLogin
)
from .ia_engine import analizar_ticket_con_ia, conversar_con_analista_ia, reclasificar_ticket_con_ia, generar_resumen_cliente


# ─────────────────────────────────────────────────────────────
# ERROR HANDLERS (404 / 500)
# ─────────────────────────────────────────────────────────────

def custom_404(request, exception):
    """Handler personalizado para errores 404."""
    return render(request, '404.html', status=404)


def custom_500(request):
    """Handler personalizado para errores 500."""
    return render(request, '500.html', status=500)




# ─────────────────────────────────────────────────────────────
# DECORADORES DE ROLES
# ─────────────────────────────────────────────────────────────

def rol_requerido(*roles):
    """Decorador para restringir el acceso a vistas basado en el rol del PerfilUsuario."""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')

            if not hasattr(request.user, 'perfil'):
                return render(request, 'tickets/acceso_denegado.html', {
                    'mensaje': 'Su usuario no tiene un perfil configurado. Contacte al administrador.'
                }, status=403)

            # Si hay simulación activa, el usuario real es superadmin → permitir siempre
            if getattr(request, 'simulacion_activa', False):
                return view_func(request, *args, **kwargs)

            if request.user.perfil.rol not in roles and 'superadmin' not in [request.user.perfil.rol]:
                return render(request, 'tickets/acceso_denegado.html', {
                    'mensaje': f'El rol {request.user.perfil.get_rol_display()} no tiene permisos para esta sección.'
                }, status=403)

            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# ─────────────────────────────────────────────────────────────
# HELPERS DE EMAIL
# ─────────────────────────────────────────────────────────────

def _enviar_acuse_recibo(ticket, request=None):
    """Envía acuse de recibo al cliente cuando se crea un ticket por email."""
    try:
        send_mail(
            subject=f'✅ Recibimos su PQRS — Caso #{ticket.ticket_id} | Unidossis',
            message=(
                f'Estimado/a {ticket.remitente_nombre or "Cliente"},\n\n'
                f'Confirmamos que hemos recibido correctamente su solicitud.\n\n'
                f'  • Número de caso: {ticket.ticket_id}\n'
                f'  • Tipo de solicitud: {ticket.get_tipo_solicitud_display()}\n'
                f'  • Asunto: {ticket.asunto}\n'
                f'  • Fecha de ingreso: {ticket.fecha_ingreso.strftime("%d/%m/%Y %H:%M")}\n\n'
                f'Nuestro equipo revisará su caso y le dará respuesta en los tiempos establecidos '
                f'según nuestra política de servicio.\n\n'
                f'Gracias por comunicarse con Unidossis.\n\n'
                f'— Servicio al Cliente UNIDOSSIS\n'
                f'  servicio.cliente@unidossis.com.co'
            ),
            from_email='Servicio al cliente Unidossis <servicio.cliente@unidossis.com.co>',
            recipient_list=[ticket.remitente_email],
            fail_silently=True,
        )
        ticket.auto_respuesta_enviada = True
        ticket.save(update_fields=['auto_respuesta_enviada'])
        LogActividad.objects.create(
            ticket=ticket, usuario=None,
            accion='Acuse de recibo enviado automáticamente al cliente',
            detalle=f'Correo enviado a: {ticket.remitente_email}'
        )
    except Exception:
        pass


def _enviar_respuesta_formal_y_csat(ticket, request=None):
    """
    Envía la respuesta formal al cierre e incluye enlace a encuesta CSAT.
    Crea el registro de encuesta con token único.
    """
    try:
        # Crear o recuperar encuesta CSAT
        encuesta, created = EncuestaSatisfaccion.objects.get_or_create(ticket=ticket)

        # Construir URL de la encuesta
        if request:
            base_url = request.build_absolute_uri('/')[:-1]
        else:
            base_url = 'https://pqrs.unidossis.com.co'

        url_encuesta = f'{base_url}/encuesta/{encuesta.token}/'

        mensaje = (
            f'Estimado/a {ticket.remitente_nombre or "Cliente"},\n\n'
            f'Nos complace informarle que hemos dado respuesta a su solicitud PQRS:\n\n'
            f'  • Número de caso: {ticket.ticket_id}\n'
            f'  • Asunto: {ticket.asunto}\n\n'
            f'RESPUESTA OFICIAL DE UNIDOSSIS:\n'
            f'{"─" * 50}\n'
            f'{ticket.respuesta_oficial or "Ver respuesta en el portal de clientes."}\n'
            f'{"─" * 50}\n\n'
            f'📊 ENCUESTA DE SATISFACCIÓN:\n'
            f'Por favor califique nuestra atención haciendo clic en el siguiente enlace:\n'
            f'{url_encuesta}\n\n'
            f'Su opinión es muy importante para nosotros y nos ayuda a mejorar continuamente.\n\n'
            f'Gracias por su confianza en Unidossis.\n\n'
            f'— Servicio al Cliente UNIDOSSIS\n'
            f'  servicio.cliente@unidossis.com.co'
        )

        send_mail(
            subject=f'📋 Respuesta a su PQRS #{ticket.ticket_id} — Unidossis',
            message=mensaje,
            from_email='Servicio al cliente Unidossis <servicio.cliente@unidossis.com.co>',
            recipient_list=[ticket.remitente_email],
            fail_silently=False,
        )
        ticket.respuesta_formal_enviada = True
        ticket.save(update_fields=['respuesta_formal_enviada'])
        LogActividad.objects.create(
            ticket=ticket, usuario=request.user if request else None,
            accion='Respuesta formal enviada al cliente con encuesta CSAT',
            detalle=f'URL encuesta: {url_encuesta}'
        )
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────
# CAMBIAR CONTRASEÑA (Autoservicio)
# ─────────────────────────────────────────────────────────────

@login_required
def cambiar_password_view(request):
    """Permite a cualquier usuario autenticado cambiar su propia contraseña."""
    if request.method != 'POST':
        return JsonResponse({'tipo': 'error', 'texto': 'Método no permitido.'}, status=405)

    current_password = request.POST.get('current_password', '')
    new_password = request.POST.get('new_password', '')
    confirm_password = request.POST.get('confirm_password', '')

    if not request.user.check_password(current_password):
        return JsonResponse({'tipo': 'error', 'texto': 'La contraseña actual es incorrecta.'})

    if len(new_password) < 8:
        return JsonResponse({'tipo': 'error', 'texto': 'La nueva contraseña debe tener al menos 8 caracteres.'})

    if new_password != confirm_password:
        return JsonResponse({'tipo': 'error', 'texto': 'Las contraseñas nuevas no coinciden.'})

    if current_password == new_password:
        return JsonResponse({'tipo': 'error', 'texto': 'La nueva contraseña debe ser diferente a la actual.'})

    request.user.set_password(new_password)
    request.user.save()
    update_session_auth_hash(request, request.user)  # Mantener la sesión activa

    # Desactivar el flag de cambio forzado de contraseña
    if hasattr(request.user, 'perfil'):
        request.user.perfil.debe_cambiar_password = False
        request.user.perfil.save(update_fields=['debe_cambiar_password'])

    return JsonResponse({'tipo': 'exito', 'texto': 'Su contraseña ha sido actualizada exitosamente.'})


# ─────────────────────────────────────────────────────────────
# DASHBOARD PRINCIPAL
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'agente')
def dashboard_view(request):
    perfil = request.user.perfil

    # Base query según rol
    if perfil.rol in ['superadmin', 'admin_pqrs', 'agente']:
        tickets_query = Ticket.objects.all()
    elif perfil.rol == 'director_regional':
        # En simulación: si no se eligió regional específica, mostrar todos (modo preview)
        if perfil.regional:
            tickets_query = Ticket.objects.filter(regional=perfil.regional)
        else:
            tickets_query = Ticket.objects.all()
    else:
        # Fallback seguro para cualquier otro rol o estado inesperado
        tickets_query = Ticket.objects.all()

    # Filtro por cliente (Para Inteligencia Analítica y buscador inteligente)
    cliente_id = request.GET.get('cliente_id')
    cliente_filtrado_nombre = None
    if cliente_id:
        tickets_query = tickets_query.filter(cliente_rel_id=cliente_id)
        try:
            cliente_filtrado_nombre = Cliente.objects.get(id=cliente_id).nombre
        except Cliente.DoesNotExist:
            cliente_filtrado_nombre = None

    # Motor de búsqueda
    search_query = request.GET.get('q')
    if search_query:
        tickets_query = tickets_query.filter(
            Q(ticket_id__icontains=search_query) |
            Q(asunto__icontains=search_query) |
            Q(remitente_nombre__icontains=search_query) |
            Q(remitente_email__icontains=search_query) |
            Q(entidad_cliente__icontains=search_query) |
            Q(institucion__icontains=search_query) |
            Q(responsable__icontains=search_query) |
            Q(cuerpo__icontains=search_query) |
            Q(cliente_rel__nombre__icontains=search_query)
        )

    # Obtener IDs de clientes que tienen tickets en esta consulta (rol + búsqueda + filtro actual, sin estado)
    clientes_con_tickets_ids = list(tickets_query.exclude(cliente_rel__isnull=True).values_list('cliente_rel_id', flat=True).distinct())

    # Estadísticas generales y totales de las pestañas (Calculados ANTES de aplicar el filtro de estado)
    total_tickets = tickets_query.count()
    total_abiertos = tickets_query.exclude(estado__in=['resuelto', 'cancelado']).count()
    total_cerrados = tickets_query.filter(estado='resuelto').count()
    total_revision = tickets_query.filter(estado='revision').count()
    total_cancelados = tickets_query.filter(estado='cancelado').count()
    cumplimiento = round((total_cerrados / total_tickets * 100), 1) if total_tickets > 0 else 0

    # SLA semáforo
    tickets_abiertos = list(tickets_query.exclude(estado__in=['resuelto', 'cancelado']))
    vencidas = sum(1 for t in tickets_abiertos if t.estado_sla() == 'vencido')
    peligro = sum(1 for t in tickets_abiertos if t.estado_sla() == 'peligro')
    bien = sum(1 for t in tickets_abiertos if t.estado_sla() == 'bien')

    # SLA health %
    total_activos_sla = vencidas + peligro + bien
    sla_health = round((bien / total_activos_sla * 100)) if total_activos_sla > 0 else 100

    # Tickets de hoy
    from datetime import timedelta
    hoy = timezone.now().date()
    # Compatibilidad con SQLite (datetime)
    tickets_hoy = sum(1 for t in tickets_query if t.fecha_ingreso.date() == hoy)

    # APLICAR EL FILTRO DE ESTADO A LA TABLA PRINCIPAL (Al final)
    filtro_estado = request.GET.get('estado')
    if filtro_estado:
        tickets_query = tickets_query.filter(estado=filtro_estado)

    # Paginación para la tabla
    paginator = Paginator(tickets_query.order_by('-fecha_ingreso'), 30)
    from django.db.models import Count
    regional_dict = dict(Ticket.REGIONAL_CHOICES)
    linea_dict = dict(Ticket.LINEA_CHOICES)

    top_regionales = list(tickets_query.exclude(regional__isnull=True).exclude(regional='').values('regional').annotate(
        total=Count('id')
    ).order_by('-total')[:4])

    for r in top_regionales:
        r['nombre'] = regional_dict.get(r['regional'], r['regional'])
        r['porcentaje'] = round((r['total'] / total_tickets * 100)) if total_tickets > 0 else 0
        r['abiertos'] = tickets_query.filter(regional=r['regional']).exclude(estado__in=['resuelto', 'cancelado']).count()

    top_lineas = list(tickets_query.exclude(linea_servicio__isnull=True).exclude(linea_servicio='').values('linea_servicio').annotate(
        total=Count('id')
    ).order_by('-total')[:4])

    for l in top_lineas:
        l['nombre'] = linea_dict.get(l['linea_servicio'], l['linea_servicio'])
        l['porcentaje'] = round((l['total'] / total_tickets * 100)) if total_tickets > 0 else 0

    clientes_disponibles = Cliente.objects.filter(id__in=clientes_con_tickets_ids, activo=True).order_by('nombre')

    context = {
        'tickets': tickets_query.order_by('-fecha_ingreso'),
        'cantidad_total': total_tickets,
        'cantidad_urgentes': total_abiertos,
        'total_cerrados': total_cerrados,
        'total_revision': total_revision,
        'total_cancelados': total_cancelados,
        'cumplimiento': cumplimiento,
        'vencidas': vencidas,
        'peligro': peligro,
        'bien': bien,
        'sla_health': sla_health,
        'tickets_hoy': tickets_hoy,
        'top_regionales': top_regionales,
        'top_lineas': top_lineas,
        'filtro_actual': filtro_estado,
        'perfil': perfil,
        'nav_active': 'dashboard',
        'clientes_disponibles': clientes_disponibles,
        'cliente_id_selected': cliente_id,
        'cliente_filtrado_nombre': cliente_filtrado_nombre,
        # Choices para modal de creación manual
        'TIPO_CHOICES': Ticket.TIPO_SOLICITUD_CHOICES,
        'REGIONAL_CHOICES': Ticket.REGIONAL_CHOICES,
    }
    return render(request, 'tickets/dashboard.html', context)


# ─────────────────────────────────────────────────────────────
# DETALLE DE TICKET
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'agente', 'cliente')
def ticket_detail_view(request, ticket_id):
    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)
    perfil = request.user.perfil

    # Calcular si el agente está asignado a este ticket (su nombre aparece en 'responsable')
    es_agente_asignado = False
    if perfil.rol == 'agente':
        responsable_str = (ticket.responsable or "").lower()
        nombre_usuario = request.user.get_full_name().lower() if request.user.get_full_name() else ""
        es_agente_asignado = (
            request.user.username.lower() in responsable_str or
            (nombre_usuario and nombre_usuario in responsable_str) or
            (request.user.last_name and request.user.last_name.lower() in responsable_str) or
            (request.user.first_name and request.user.first_name.lower() in responsable_str)
        )

    # Validaciones de seguridad por rol
    if perfil.rol == 'cliente':
        if ticket.cliente_rel != perfil.cliente:
            return redirect('acceso_denegado')
    elif perfil.rol == 'director_regional':
        if ticket.regional != perfil.regional:
            return redirect('acceso_denegado')

    adjuntos_cliente = ticket.archivos_adjuntos.filter(es_respuesta_agente=False, es_soporte_interno=False)
    adjuntos_unidossis = ticket.archivos_adjuntos.filter(es_respuesta_agente=True)
    adjuntos_internos = ticket.archivos_adjuntos.filter(es_soporte_interno=True)
    comentarios = ticket.comentarios.all()
    logs = ticket.logs.all()[:20]

    if request.method == 'POST' and perfil.rol != 'cliente':
        # Los agentes solo pueden hacer POST si están asignados al ticket
        if perfil.rol == 'agente' and not es_agente_asignado:
            return redirect('ticket_detail', ticket_id=ticket.ticket_id)
        estado_anterior = ticket.estado
        nuevo_estado = request.POST.get('nuevo_estado')
        respuesta_texto = request.POST.get('respuesta_oficial')

        if nuevo_estado in dict(Ticket.STATUS_CHOICES):
            ticket.estado = nuevo_estado

        proceso = request.POST.get('proceso')
        linea = request.POST.get('linea_servicio')
        tipificacion = request.POST.get('tipificacion')
        criticidad = request.POST.get('criticidad')
        regional = request.POST.get('regional')

        if proceso: ticket.proceso = proceso
        if linea: ticket.linea_servicio = linea
        if tipificacion: ticket.tipificacion = tipificacion
        if criticidad: ticket.criticidad = criticidad

        # Regional: superadmin, admin_pqrs y agentes asignados pueden cambiarla
        if regional is not None and perfil.rol in ('superadmin', 'admin_pqrs', 'agente'):
            regional_anterior = ticket.regional
            ticket.regional = regional
            if regional != regional_anterior:
                LogActividad.objects.create(
                    ticket=ticket, usuario=request.user,
                    accion=f'Regional cambiada: {regional_anterior or "N/A"} → {regional or "N/A"}',
                    detalle=f'Cambiado por: {request.user.get_full_name() or request.user.username}'
                )

        responsable_manual = request.POST.get('responsable')
        if responsable_manual: ticket.responsable = responsable_manual

        if respuesta_texto is not None and perfil.rol != 'agente':
            ticket.respuesta_oficial = respuesta_texto

        ticket.save()

        # Log de actividad automático
        if nuevo_estado and nuevo_estado != estado_anterior:
            LogActividad.objects.create(
                ticket=ticket, usuario=request.user,
                accion=f'Estado cambiado: {estado_anterior} → {nuevo_estado}',
                detalle=f'Cambiado por: {request.user.get_full_name() or request.user.username}'
            )

        # Archivos adjuntos del agente (respuesta al cliente)
        if request.FILES.getlist('archivos_agente'):
            for archivo_subido in request.FILES.getlist('archivos_agente'):
                ArchivoAdjunto.objects.create(
                    ticket=ticket,
                    archivo=archivo_subido,
                    es_respuesta_agente=True,
                    subido_por_sistema=False
                )

        # Soportes y evidencias internas (no visibles para el cliente)
        if request.FILES.getlist('soportes_internos'):
            for archivo_subido in request.FILES.getlist('soportes_internos'):
                ArchivoAdjunto.objects.create(
                    ticket=ticket,
                    archivo=archivo_subido,
                    es_soporte_interno=True,
                    subido_por_sistema=False
                )

        # Cerrar y enviar respuesta formal + CSAT (no permitido para agentes)
        if 'cerrar_y_enviar' in request.POST and perfil.rol != 'agente':
            ticket.estado = 'resuelto'
            ticket.save()
            _enviar_respuesta_formal_y_csat(ticket, request)

        return redirect('ticket_detail', ticket_id=ticket.ticket_id)

    template_name = 'tickets/detalle.html'
    if perfil.rol == 'cliente':
        template_name = 'tickets/cliente/detalle.html'

    return render(request, template_name, {
        'ticket': ticket,
        'adjuntos_cliente': adjuntos_cliente,
        'adjuntos_unidossis': adjuntos_unidossis,
        'adjuntos_internos': adjuntos_internos,
        'comentarios': comentarios,
        'logs': logs,
        'STATUS_CHOICES': Ticket.STATUS_CHOICES,
        'AREA_CHOICES': Ticket.AREA_CHOICES,
        'LINEA_CHOICES': Ticket.LINEA_CHOICES,
        'TIPIFICACION_CHOICES': Ticket.TIPIFICACION_CHOICES,
        'CRITICIDAD_CHOICES': Ticket.CRITICIDAD_CHOICES,
        'REGIONAL_CHOICES': Ticket.REGIONAL_CHOICES,
        'perfil': perfil,
        'nav_active': 'dashboard',
        'cliente': perfil.cliente if perfil.rol == 'cliente' else None,
        'es_agente_asignado': es_agente_asignado,  # Pasado al template para controlar edición
    })


# ─────────────────────────────────────────────────────────────
# PORTAL PÚBLICO DE CLIENTES
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('cliente')
def public_pqrs_view(request):
    # ── Si hay simulación activa de cliente, usar el mock del middleware ──
    if getattr(request, 'simulacion_activa', False) and getattr(request, 'cliente_simulado', None):
        perfil_cliente = request.cliente_simulado
    else:
        perfil_cliente = getattr(request.user, 'cliente_perfil', None)

    simulacion_activa = getattr(request, 'simulacion_activa', False)

    if perfil_cliente and not perfil_cliente.activo and not simulacion_activa:
        return render(request, 'tickets/acceso_denegado.html', {
            'mensaje': 'Su cuenta institucional se encuentra inactiva. Comuníquese con el equipo de Unidossis para más información.'
        }, status=403)

    ctx_base = {
        'TIPO_CHOICES': Ticket.TIPO_SOLICITUD_CHOICES,
        'perfil_cliente': perfil_cliente,
        'simulacion_activa': simulacion_activa,
        'rol_simulado': getattr(request, 'rol_simulado', None),
    }

    if request.method == 'POST':
        institucion = request.POST.get('institucion')
        entidad = institucion
        ciudad = request.POST.get('ciudad')
        nombre = request.POST.get('nombre')
        cargo = request.POST.get('cargo')
        telefono = request.POST.get('telefono')
        email = request.POST.get('email')
        tipo = request.POST.get('tipo_solicitud')
        asunto = request.POST.get('asunto')
        descripcion = request.POST.get('descripcion')

        if not (institucion and nombre and email and asunto and descripcion):
            ctx_base['error'] = 'Por favor, complete todos los campos obligatorios (Institución, Nombre, Correo, Asunto y Descripción).'
            return render(request, 'tickets/portal_form.html', ctx_base)

        # En simulación no se crea ticket real, solo se muestra éxito
        if simulacion_activa:
            from types import SimpleNamespace
            ticket_fake = SimpleNamespace(
                ticket_id='SIM-PREVIEW',
                remitente_email=email,
                asunto=asunto,
            )
            return render(request, 'tickets/portal_success.html', {
                'ticket': ticket_fake,
                'simulacion_activa': True,
            })

        # Procesamiento por IA
        analisis = analizar_ticket_con_ia(asunto, descripcion)

        nuevo_ticket = Ticket.objects.create(
            cliente_rel=perfil_cliente,
            entidad_cliente=entidad,
            institucion=institucion,
            ciudad=ciudad,
            remitente_nombre=nombre,
            solicitante_cargo=cargo,
            telefono=telefono,
            remitente_email=email,
            tipo_solicitud=tipo,
            asunto=asunto,
            cuerpo=descripcion,
            regional=perfil_cliente.regional if (perfil_cliente and perfil_cliente.regional) else 'liquidos',
            proceso=analisis['proceso'],
            linea_servicio=analisis['linea'],
            tipificacion=analisis['tipificacion'],
            criticidad=analisis['criticidad'],
            analisis_ia=analisis['analisis_ia'],
            clasificado_por_ia=True
        )

        # Archivos adjuntos
        if request.FILES.getlist('adjuntos'):
            for f in request.FILES.getlist('adjuntos'):
                ArchivoAdjunto.objects.create(ticket=nuevo_ticket, archivo=f)

        # Log de actividad
        LogActividad.objects.create(
            ticket=nuevo_ticket, usuario=request.user,
            accion='Ticket creado por el cliente desde el portal',
            detalle=f'Clasificado por IA: {analisis["tipificacion"]} / {analisis["criticidad"]}'
        )

        return render(request, 'tickets/portal_success.html', {'ticket': nuevo_ticket})

    return render(request, 'tickets/portal_form.html', ctx_base)


# ─────────────────────────────────────────────────────────────
# LOGIN / LOGOUT
# ─────────────────────────────────────────────────────────────

def _get_client_ip(request):
    """Obtiene la IP real del cliente (soporta proxies)."""
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded:
        return x_forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '0.0.0.0')


def login_view(request):
    MAX_INTENTOS = 5
    MINUTOS_BLOQUEO = 15

    if request.method == 'POST':
        ip = _get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
        username_input = request.POST.get('username', '')

        # ── Rate limiting: verificar intentos recientes ──
        desde = timezone.now() - timezone.timedelta(minutes=MINUTOS_BLOQUEO)
        intentos_fallidos = IntentoLogin.objects.filter(
            ip=ip, exitoso=False, fecha__gte=desde
        ).count()

        if intentos_fallidos >= MAX_INTENTOS:
            # Registrar intento bloqueado
            IntentoLogin.objects.create(ip=ip, username=username_input, exitoso=False, user_agent=user_agent)
            return render(request, 'tickets/login.html', {
                'form': AuthenticationForm(),
                'error_bloqueo': f'Demasiados intentos fallidos. Intente de nuevo en {MINUTOS_BLOQUEO} minutos.'
            })

        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # Registrar login exitoso
            IntentoLogin.objects.create(ip=ip, username=user.username, exitoso=True, user_agent=user_agent)

            # Log de auditoría
            LogActividad.objects.create(
                ticket=None, usuario=user,
                accion='Inicio de sesión',
                detalle=f'IP: {ip} | Navegador: {user_agent[:100]}'
            )

            try:
                perfil = user.perfil

                # ── Cambio forzado de contraseña ──
                if perfil.debe_cambiar_password:
                    return render(request, 'tickets/cambiar_password_obligatorio.html', {
                        'perfil': perfil,
                    })

                if perfil.rol == 'cliente':
                    return redirect('portal_cliente_dashboard')
                else:
                    # Superadmin, admin, director, agente → siempre al dashboard
                    return redirect('dashboard')
            except PerfilUsuario.DoesNotExist:
                return redirect('dashboard')
        else:
            # Registrar intento fallido
            IntentoLogin.objects.create(ip=ip, username=username_input, exitoso=False, user_agent=user_agent)
    else:
        form = AuthenticationForm()
    return render(request, 'tickets/login.html', {'form': form})


def logout_view(request):
    # Log de auditoría antes del logout
    if request.user.is_authenticated:
        ip = _get_client_ip(request)
        LogActividad.objects.create(
            ticket=None, usuario=request.user,
            accion='Cierre de sesión',
            detalle=f'IP: {ip}'
        )
    logout(request)
    return redirect('login')


# ─────────────────────────────────────────────────────────────
# GESTIÓN DE CLIENTES
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def gestionar_clientes_view(request):
    """Panel para que el admin asigne usuario y contraseña a los clientes."""
    mensaje = None

    if request.method == 'POST':
        action = request.POST.get('action')
        is_ajax = request.POST.get('is_ajax') == '1'

        if action == 'crear':
            cliente_id = request.POST.get('cliente_id')
            username = request.POST.get('username')
            password = request.POST.get('password')
            email = request.POST.get('email', '')

            if cliente_id and username and password:
                try:
                    cliente = Cliente.objects.get(id=cliente_id)
                    if User.objects.filter(username=username).exists():
                        mensaje = {'tipo': 'error', 'texto': f'El usuario "{username}" ya existe.'}
                    else:
                        nuevo_user = User.objects.create_user(username=username, password=password, email=email)
                        nuevo_user.first_name = cliente.nombre[:30]
                        nuevo_user.save()
                        cliente.user = nuevo_user
                        PerfilUsuario.objects.get_or_create(
                            user=nuevo_user,
                            defaults={'rol': 'cliente', 'cliente': cliente}
                        )
                        nueva_regional = request.POST.get('regional')
                        nuevo_email_p = request.POST.get('email_principal')
                        nuevo_email_a = request.POST.get('emails_adicionales')
                        if nueva_regional: cliente.regional = nueva_regional
                        if nuevo_email_p: cliente.email_principal = nuevo_email_p
                        if nuevo_email_a: cliente.emails_adicionales = nuevo_email_a
                        cliente.save()
                        mensaje = {'tipo': 'exito', 'texto': f'Acceso creado para {cliente.nombre}: usuario "{username}".'}
                except Cliente.DoesNotExist:
                    mensaje = {'tipo': 'error', 'texto': 'Cliente no encontrado.'}

        elif action == 'cambiar_password':
            user_id = request.POST.get('user_id')
            new_password = request.POST.get('new_password')
            if user_id and new_password:
                try:
                    user_obj = User.objects.get(id=user_id)
                    user_obj.set_password(new_password)
                    user_obj.save()
                    mensaje = {'tipo': 'exito', 'texto': f'Contraseña actualizada para {user_obj.username}.'}
                except User.DoesNotExist:
                    mensaje = {'tipo': 'error', 'texto': 'Usuario no encontrado.'}

        elif action == 'crear_institucion':
            nombre_institucion = request.POST.get('nombre_institucion', '').strip()
            ciudad_id = request.POST.get('ciudad_id')
            regional = request.POST.get('regional')
            email_p = request.POST.get('email_principal', '').strip()
            emails_a = request.POST.get('emails_adicionales', '').strip()
            if nombre_institucion and email_p:
                if Cliente.objects.filter(nombre__iexact=nombre_institucion).exists():
                    mensaje = {'tipo': 'error', 'texto': f'La institución "{nombre_institucion}" ya existe.'}
                else:
                    nueva_inst = Cliente.objects.create(
                        nombre=nombre_institucion, email_principal=email_p,
                        emails_adicionales=emails_a, regional=regional
                    )
                    MaestroInstitucion.objects.get_or_create(nombre=nombre_institucion)
                    if ciudad_id:
                        try:
                            nueva_inst.ciudad = Ciudad.objects.get(id=ciudad_id)
                            nueva_inst.save()
                        except Ciudad.DoesNotExist:
                            pass
                    mensaje = {'tipo': 'exito', 'texto': f'"{nueva_inst.nombre}" ha sido creada exitosamente.'}
            else:
                mensaje = {'tipo': 'error', 'texto': 'El nombre y el correo principal son obligatorios.'}

        elif action == 'inactivar':
            cliente_id = request.POST.get('cliente_id')
            try:
                cliente = Cliente.objects.get(id=cliente_id)
                cliente.activo = not cliente.activo
                cliente.save()
                estado_txt = 'activado' if cliente.activo else 'marcado como inactivo'
                mensaje = {'tipo': 'exito', 'texto': f'"{cliente.nombre}" ha sido {estado_txt} exitosamente.'}
            except Cliente.DoesNotExist:
                mensaje = {'tipo': 'error', 'texto': 'Cliente no encontrado.'}

        elif action == 'editar_cliente':
            cliente_id = request.POST.get('cliente_id')
            nombre = request.POST.get('nombre', '').strip()
            regional = request.POST.get('regional')
            email_p = request.POST.get('email_principal', '').strip()
            emails_a = request.POST.get('emails_adicionales', '').strip()
            ciudad_id = request.POST.get('ciudad_id')
            activo = request.POST.get('activo') == 'on'
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            try:
                cliente = Cliente.objects.get(id=cliente_id)
                cliente.nombre = nombre
                cliente.regional = regional
                cliente.email_principal = email_p
                cliente.emails_adicionales = emails_a
                cliente.activo = activo
                if ciudad_id:
                    try:
                        cliente.ciudad = Ciudad.objects.get(id=ciudad_id)
                    except Ciudad.DoesNotExist:
                        pass
                if username and not cliente.user:
                    if User.objects.filter(username=username).exists():
                        mensaje = {'tipo': 'error', 'texto': f'El usuario "{username}" ya está en uso.'}
                    else:
                        nuevo_user = User.objects.create_user(username=username, password=password or '12345', email=email_p)
                        nuevo_user.first_name = cliente.nombre[:30]
                        nuevo_user.save()
                        PerfilUsuario.objects.get_or_create(
                            user=nuevo_user, defaults={'rol': 'cliente', 'cliente': cliente}
                        )
                        cliente.user = nuevo_user
                elif cliente.user and password:
                    cliente.user.set_password(password)
                    cliente.user.save()
                cliente.save()
                mensaje = {'tipo': 'exito', 'texto': f'Configuración de "{cliente.nombre}" actualizada con éxito.'}
            except Cliente.DoesNotExist:
                mensaje = {'tipo': 'error', 'texto': 'Cliente no encontrado.'}

        elif action == 'eliminar':
            cliente_id = request.POST.get('cliente_id')
            try:
                cliente = Cliente.objects.get(id=cliente_id)
                nombre_eliminado = cliente.nombre
                if cliente.user:
                    cliente.user.delete()
                else:
                    cliente.delete()
                mensaje = {'tipo': 'exito', 'texto': f'"{nombre_eliminado}" ha sido eliminado permanentemente del sistema.'}
            except Cliente.DoesNotExist:
                mensaje = {'tipo': 'error', 'texto': 'Cliente no encontrado.'}

        # ─── Acciones de Usuarios Internos ────────────────────
        elif action == 'crear_usuario':
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            rol = request.POST.get('rol', '')
            regional = request.POST.get('regional', '') or None

            if username and password and rol:
                if User.objects.filter(username=username).exists():
                    mensaje = {'tipo': 'error', 'texto': f'El usuario "{username}" ya existe en el sistema.'}
                else:
                    nuevo_user = User.objects.create_user(
                        username=username, password=password,
                        email=email, first_name=first_name, last_name=last_name
                    )
                    nuevo_user.is_staff = True
                    nuevo_user.save()
                    telefono = request.POST.get('telefono', '').strip()
                    PerfilUsuario.objects.create(user=nuevo_user, rol=rol, regional=regional, telefono=telefono or None)
                    rol_display = dict(PerfilUsuario.ROL_CHOICES).get(rol, rol)
                    mensaje = {'tipo': 'exito', 'texto': f'Usuario "{username}" creado exitosamente como {rol_display}.'}
            else:
                mensaje = {'tipo': 'error', 'texto': 'Nombre de usuario, contraseña y rol son obligatorios.'}

        elif action == 'editar_usuario':
            user_id = request.POST.get('user_id')
            try:
                user_obj = User.objects.get(id=user_id)
                if hasattr(user_obj, 'perfil') and user_obj.perfil.rol == 'superadmin':
                    mensaje = {'tipo': 'error', 'texto': 'No se puede editar un Super Administrador desde esta consola.'}
                else:
                    new_username = request.POST.get('username', '').strip()
                    # Validar username único (si cambió)
                    if new_username and new_username != user_obj.username:
                        if User.objects.filter(username=new_username).exclude(id=user_obj.id).exists():
                            mensaje = {'tipo': 'error', 'texto': f'El usuario "{new_username}" ya existe en el sistema.'}
                        else:
                            user_obj.username = new_username
                    if not mensaje:  # Solo continuar si no hubo error
                        user_obj.first_name = request.POST.get('first_name', '').strip()
                        user_obj.last_name = request.POST.get('last_name', '').strip()
                        user_obj.email = request.POST.get('email', '').strip()
                        new_pass = request.POST.get('password', '').strip()
                        if new_pass:
                            user_obj.set_password(new_pass)
                        user_obj.save()
                        perfil_obj = user_obj.perfil
                        perfil_obj.rol = request.POST.get('rol', perfil_obj.rol)
                        perfil_obj.regional = request.POST.get('regional', '') or None
                        perfil_obj.telefono = request.POST.get('telefono', '').strip() or None
                        perfil_obj.save()
                        mensaje = {'tipo': 'exito', 'texto': f'Usuario "{user_obj.username}" actualizado correctamente.'}
            except User.DoesNotExist:
                mensaje = {'tipo': 'error', 'texto': 'Usuario no encontrado.'}

        elif action == 'toggle_usuario':
            user_id = request.POST.get('user_id')
            try:
                user_obj = User.objects.get(id=user_id)
                if hasattr(user_obj, 'perfil') and user_obj.perfil.rol == 'superadmin':
                    mensaje = {'tipo': 'error', 'texto': 'No se puede desactivar un Super Administrador.'}
                else:
                    user_obj.is_active = not user_obj.is_active
                    user_obj.save()
                    estado_txt = 'activado' if user_obj.is_active else 'desactivado'
                    mensaje = {'tipo': 'exito', 'texto': f'Usuario "{user_obj.username}" ha sido {estado_txt}.'}
            except User.DoesNotExist:
                mensaje = {'tipo': 'error', 'texto': 'Usuario no encontrado.'}

        elif action == 'eliminar_usuario':
            user_id = request.POST.get('user_id')
            try:
                user_obj = User.objects.get(id=user_id)
                if hasattr(user_obj, 'perfil') and user_obj.perfil.rol == 'superadmin':
                    mensaje = {'tipo': 'error', 'texto': 'No se puede eliminar un Super Administrador.'}
                else:
                    nombre_usr = user_obj.get_full_name() or user_obj.username
                    user_obj.delete()
                    mensaje = {'tipo': 'exito', 'texto': f'Usuario "{nombre_usr}" eliminado permanentemente del sistema.'}
            except User.DoesNotExist:
                mensaje = {'tipo': 'error', 'texto': 'Usuario no encontrado.'}

        if is_ajax and mensaje:
            return JsonResponse(mensaje)

    clientes_sin_acceso = Cliente.objects.filter(user__isnull=True)[:50]
    todos_los_clientes = Cliente.objects.all().select_related('user', 'ciudad')

    # Usuarios internos (excluye clientes y superadmins)
    usuarios_internos = PerfilUsuario.objects.filter(
        rol__in=['admin_pqrs', 'director_regional', 'agente']
    ).select_related('user').order_by('user__first_name')

    ROL_CHOICES_INTERNOS = [
        ('admin_pqrs', 'Administrador PQRS'),
        ('director_regional', 'Director Regional'),
        ('agente', 'Agente / Consultor'),
    ]

    tab_activa = request.POST.get('_tab', request.GET.get('tab', 'clientes'))

    return render(request, 'tickets/gestionar_clientes.html', {
        'clientes_sin_acceso': clientes_sin_acceso,
        'todos_los_clientes': todos_los_clientes,
        'usuarios_internos': usuarios_internos,
        'mensaje': mensaje,
        'REGIONAL_CHOICES': Cliente.REGIONAL_CHOICES,
        'ROL_CHOICES_INTERNOS': ROL_CHOICES_INTERNOS,
        'perfil': request.user.perfil,
        'nav_active': 'clientes',
        'tab_activa': tab_activa,
    })


# ─────────────────────────────────────────────────────────────
# PORTAL CLIENTE DASHBOARD
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('cliente')
def portal_cliente_dashboard(request):
    """Dashboard para que el cliente vea sus propios tickets."""
    from django.db.models import Count
    from django.utils import timezone
    import json as _json

    perfil = request.user.perfil

    # ── En modo simulación, usar el cliente inyectado por el middleware ──
    simulacion_activa = getattr(request, 'simulacion_activa', False)
    if simulacion_activa and getattr(request, 'cliente_simulado', None):
        cliente = request.cliente_simulado
    else:
        cliente = perfil.cliente

    if not cliente:
        return render(request, 'tickets/acceso_denegado.html', {
            'mensaje': 'Su usuario no está vinculado a ninguna institución clínica.'
        })

    tickets_qs = Ticket.objects.filter(cliente_rel=cliente).order_by('-fecha_ingreso')

    # ── KPIs principales ──
    total_tickets       = tickets_qs.count()
    tickets_resueltos   = tickets_qs.filter(estado='resuelto').count()
    tickets_cancelados  = tickets_qs.filter(estado='cancelado').count()
    tickets_en_proceso  = tickets_qs.exclude(estado__in=['resuelto', 'cancelado']).count()

    # Vencidos SLA: abiertos con más de 14 días
    limite_sla = timezone.now() - timezone.timedelta(days=14)
    tickets_vencidos_sla = tickets_qs.exclude(
        estado__in=['resuelto', 'cancelado']
    ).filter(fecha_ingreso__lt=limite_sla).count()

    pct_cumplimiento = round((tickets_resueltos / total_tickets * 100), 1) if total_tickets else 0
    pct_en_proceso   = round((tickets_en_proceso / total_tickets * 100), 1) if total_tickets else 0

    # ── Por Tipo de Solicitud ──
    tipo_labels_map = dict(Ticket.TIPO_SOLICITUD_CHOICES)
    por_tipo_qs = (
        tickets_qs.values('tipo_solicitud')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    por_tipo = [
        {'key': r['tipo_solicitud'], 'label': tipo_labels_map.get(r['tipo_solicitud'], r['tipo_solicitud']), 'cnt': r['cnt']}
        for r in por_tipo_qs
    ]

    # ── Por Línea de Servicio (Top 6) ──
    linea_labels_map = dict(Ticket.LINEA_CHOICES)
    por_linea_qs = (
        tickets_qs.exclude(linea_servicio__isnull=True)
        .values('linea_servicio')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:6]
    )
    por_linea = [
        {'key': r['linea_servicio'], 'label': linea_labels_map.get(r['linea_servicio'], r['linea_servicio']), 'cnt': r['cnt']}
        for r in por_linea_qs
    ]

    # ── Por Criticidad ──
    crit_labels_map = {'critica': 'Crítica', 'mayor': 'Mayor', 'menor': 'Menor', 'informativa': 'Informativa'}
    crit_colors = {'critica': '#ef4444', 'mayor': '#f97316', 'menor': '#eab308', 'informativa': '#22c55e'}
    por_criticidad_qs = (
        tickets_qs.exclude(criticidad__isnull=True)
        .values('criticidad')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    por_criticidad = [
        {'key': r['criticidad'], 'label': crit_labels_map.get(r['criticidad'], r['criticidad']),
         'cnt': r['cnt'], 'color': crit_colors.get(r['criticidad'], '#6366f1')}
        for r in por_criticidad_qs
    ]

    # ── Tendencia mensual (últimos 6 meses) ──
    hoy = timezone.now()
    tendencia = []
    for i in range(5, -1, -1):
        mes_inicio = (hoy - timezone.timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        mes_fin = (mes_inicio + timezone.timedelta(days=32)).replace(day=1)
        cnt = tickets_qs.filter(fecha_ingreso__gte=mes_inicio, fecha_ingreso__lt=mes_fin).count()
        tendencia.append({'mes': mes_inicio.strftime('%b %Y'), 'cnt': cnt})

    # ── Serializar datos para Chart.js ──
    chart_tipo   = _json.dumps({'labels': [t['label'] for t in por_tipo],   'data': [t['cnt'] for t in por_tipo]})
    chart_linea  = _json.dumps({'labels': [l['label'] for l in por_linea],  'data': [l['cnt'] for l in por_linea]})
    chart_trend  = _json.dumps({'labels': [t['mes'] for t in tendencia],    'data': [t['cnt'] for t in tendencia]})

    # ── Últimos 8 tickets para tabla ──
    tickets_recientes = list(tickets_qs[:8])

    # ── Lazy backfill de resumen IA (máx 5 por carga, corre siempre) ──
    sin_resumen = [t for t in tickets_recientes if not t.resumen_cliente_ia and t.cuerpo][:5]
    for t in sin_resumen:
        try:
            resumen = generar_resumen_cliente(t.asunto, t.cuerpo)
            if resumen:
                Ticket.objects.filter(pk=t.pk).update(resumen_cliente_ia=resumen)
                t.resumen_cliente_ia = resumen
        except Exception:
            pass

    return render(request, 'tickets/cliente/dashboard.html', {
        'cliente': cliente,
        'tickets': tickets_recientes,
        'tickets_todos': tickets_qs,
        # KPIs
        'total_tickets': total_tickets,
        'tickets_resueltos': tickets_resueltos,
        'tickets_en_proceso': tickets_en_proceso,
        'tickets_cancelados': tickets_cancelados,
        'tickets_vencidos_sla': tickets_vencidos_sla,
        'pct_cumplimiento': pct_cumplimiento,
        'pct_en_proceso': pct_en_proceso,
        # Analytics
        'por_tipo': por_tipo,
        'por_linea': por_linea,
        'por_criticidad': por_criticidad,
        # Charts JSON
        'chart_tipo': chart_tipo,
        'chart_linea': chart_linea,
        'chart_trend': chart_trend,
        # Nav
        'nav_active': 'mis_pqrs',
        'simulacion_activa': simulacion_activa,
        'rol_simulado': getattr(request, 'rol_simulado', None),
    })



# ─────────────────────────────────────────────────────────────
# ACCESO DENEGADO
# ─────────────────────────────────────────────────────────────

def acceso_denegado_view(request):
    """Vista genérica para errores de permisos."""
    mensaje = request.GET.get('mensaje', 'No tiene permisos para acceder a esta sección.')
    return render(request, 'tickets/acceso_denegado.html', {'mensaje': mensaje})


# ─────────────────────────────────────────────────────────────
# PORTAL CLIENTE: ANALÍTICAS
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('cliente')
def portal_cliente_analytics(request):
    """Página de analíticas e indicadores para el portal cliente."""
    from django.db.models import Count
    from django.utils import timezone
    import json as _json

    perfil = request.user.perfil
    simulacion_activa = getattr(request, 'simulacion_activa', False)
    if simulacion_activa and getattr(request, 'cliente_simulado', None):
        cliente = request.cliente_simulado
    else:
        cliente = perfil.cliente

    if not cliente:
        return render(request, 'tickets/acceso_denegado.html', {
            'mensaje': 'Su usuario no está vinculado a ninguna institución clínica.'
        })

    tickets_qs = Ticket.objects.filter(cliente_rel=cliente)
    total_tickets = tickets_qs.count()
    tickets_resueltos = tickets_qs.filter(estado='resuelto').count()
    pct_cumplimiento = round((tickets_resueltos / total_tickets * 100), 1) if total_tickets else 0

    # Por Tipo
    tipo_labels_map = dict(Ticket.TIPO_SOLICITUD_CHOICES)
    por_tipo = [
        {'label': tipo_labels_map.get(r['tipo_solicitud'], r['tipo_solicitud']), 'cnt': r['cnt']}
        for r in tickets_qs.values('tipo_solicitud').annotate(cnt=Count('id')).order_by('-cnt')
    ]

    # Por Línea (Top 6)
    linea_labels_map = dict(Ticket.LINEA_CHOICES)
    por_linea = [
        {'label': linea_labels_map.get(r['linea_servicio'], r['linea_servicio']), 'cnt': r['cnt']}
        for r in tickets_qs.exclude(linea_servicio__isnull=True)
            .values('linea_servicio').annotate(cnt=Count('id')).order_by('-cnt')[:6]
    ]

    # Por Criticidad
    crit_labels_map = {'critica': 'Crítica', 'mayor': 'Mayor', 'menor': 'Menor', 'informativa': 'Informativa'}
    crit_colors = {'critica': '#ef4444', 'mayor': '#f97316', 'menor': '#eab308', 'informativa': '#22c55e'}
    por_criticidad = [
        {'label': crit_labels_map.get(r['criticidad'], r['criticidad']),
         'cnt': r['cnt'], 'color': crit_colors.get(r['criticidad'], '#6366f1')}
        for r in tickets_qs.exclude(criticidad__isnull=True)
            .values('criticidad').annotate(cnt=Count('id')).order_by('-cnt')
    ]

    # Por Estado
    estado_labels = {'abierto': 'Abierto', 'revision': 'En Revisión', 'resuelto': 'Resuelto', 'cancelado': 'Cancelado'}
    por_estado = [
        {'label': estado_labels.get(r['estado'], r['estado']), 'cnt': r['cnt']}
        for r in tickets_qs.values('estado').annotate(cnt=Count('id')).order_by('-cnt')
    ]

    # Tendencia mensual (6 meses)
    hoy = timezone.now()
    tendencia = []
    for i in range(5, -1, -1):
        mes_inicio = (hoy - timezone.timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        mes_fin = (mes_inicio + timezone.timedelta(days=32)).replace(day=1)
        cnt = tickets_qs.filter(fecha_ingreso__gte=mes_inicio, fecha_ingreso__lt=mes_fin).count()
        tendencia.append({'mes': mes_inicio.strftime('%b %Y'), 'cnt': cnt})

    chart_tipo   = _json.dumps({'labels': [t['label'] for t in por_tipo],  'data': [t['cnt'] for t in por_tipo]})
    chart_linea  = _json.dumps({'labels': [l['label'] for l in por_linea], 'data': [l['cnt'] for l in por_linea]})
    chart_trend  = _json.dumps({'labels': [t['mes'] for t in tendencia],   'data': [t['cnt'] for t in tendencia]})
    chart_estado = _json.dumps({'labels': [e['label'] for e in por_estado],'data': [e['cnt'] for e in por_estado]})

    return render(request, 'tickets/cliente/analytics.html', {
        'cliente': cliente,
        'total_tickets': total_tickets,
        'tickets_resueltos': tickets_resueltos,
        'pct_cumplimiento': pct_cumplimiento,
        'por_tipo': por_tipo,
        'por_linea': por_linea,
        'por_criticidad': por_criticidad,
        'por_estado': por_estado,
        'chart_tipo': chart_tipo,
        'chart_linea': chart_linea,
        'chart_trend': chart_trend,
        'chart_estado': chart_estado,
        'nav_active': 'analiticas',
        'simulacion_activa': simulacion_activa,
        'rol_simulado': getattr(request, 'rol_simulado', None),
    })

# ─────────────────────────────────────────────────────────────
# APIs AJAX
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional')
def api_chat_analitico(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        pregunta = data.get('message', '')
        tickets = Ticket.objects.all().values(
            'ticket_id', 'remitente_nombre', 'asunto', 'estado',
            'criticidad', 'regional', 'proceso', 'linea_servicio'
        )
        contexto = list(tickets)
        respuesta = conversar_con_analista_ia(pregunta, contexto)
        return JsonResponse({'reply': respuesta})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'supervisor')
def api_buscar_clientes(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 3:
        return JsonResponse({'results': []})
    resultados = Cliente.objects.filter(nombre__icontains=query, activo=True).select_related('ciudad')[:15]
    data = [{
        'nombre': c.nombre,
        'regional': c.regional or '',
        'ciudad': c.ciudad.nombre if c.ciudad else '',
        'email': c.email_principal or '',
    } for c in resultados]
    return JsonResponse({'results': data})



@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'supervisor')
def api_buscar_ciudades(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 3:
        return JsonResponse({'results': []})
    resultados = Ciudad.objects.filter(nombre__icontains=query)[:15]
    data = [{'id': c.id, 'nombre': c.nombre} for c in resultados]
    return JsonResponse({'results': data})


@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'supervisor')
def api_buscar_cargos(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})
    resultados = Cargo.objects.filter(nombre__icontains=query)[:10]
    data = [{'nombre': c.nombre} for c in resultados]
    return JsonResponse({'results': data})


@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def api_buscar_clientes_admin(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})
    resultados = Cliente.objects.filter(nombre__icontains=query, user__isnull=True)[:15]
    data = [{'id': c.id, 'nombre': c.nombre, 'regional': c.regional,
             'email': c.email_principal, 'emails_adicionales': c.emails_adicionales or ""}
            for c in resultados]
    return JsonResponse({'results': data})


@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def api_buscar_maestro_instituciones(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})
    resultados = MaestroInstitucion.objects.filter(nombre__icontains=query)[:15]
    data = [{'id': m.id, 'nombre': m.nombre} for m in resultados]
    return JsonResponse({'results': data})


# ─────────────────────────────────────────────────────────────
# COMENTARIOS POR TICKET (AJAX)
# ─────────────────────────────────────────────────────────────

@login_required
def api_agregar_comentario(request, ticket_id):
    """Agrega un comentario al ticket. Acceso para agentes (no clientes)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    perfil = request.user.perfil
    if perfil.rol == 'cliente':
        return JsonResponse({'error': 'Sin permisos'}, status=403)

    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)
    data = json.loads(request.body)
    texto = data.get('texto', '').strip()
    visibilidad = data.get('visibilidad', 'interno')

    if not texto:
        return JsonResponse({'error': 'El comentario no puede estar vacío'}, status=400)

    comentario = ComentarioTicket.objects.create(
        ticket=ticket,
        autor=request.user,
        texto=texto,
        visibilidad=visibilidad
    )

    LogActividad.objects.create(
        ticket=ticket, usuario=request.user,
        accion=f'Comentario {"interno" if visibilidad == "interno" else "público"} agregado',
        detalle=texto[:100]
    )

    return JsonResponse({
        'ok': True,
        'comentario': {
            'id': comentario.id,
            'texto': comentario.texto,
            'autor': request.user.get_full_name() or request.user.username,
            'visibilidad': comentario.visibilidad,
            'fecha': comentario.fecha.strftime('%d/%m/%Y %H:%M'),
        }
    })


# ─────────────────────────────────────────────────────────────
# RECLASIFICACIÓN IA (AJAX)
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'supervisor')
def api_reclasificar_ia(request, ticket_id):
    """Reclasifica un ticket usando la IA con aprendizaje. Devuelve sugerencia sin aplicar."""
    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)
    resultado = reclasificar_ticket_con_ia(ticket)

    if resultado:
        return JsonResponse({'ok': True, 'clasificacion': resultado})
    return JsonResponse({'ok': False, 'error': 'IA no disponible en este momento'}, status=503)


@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional', 'supervisor')
def api_aplicar_reclasificacion(request, ticket_id):
    """Aplica la reclasificación y guarda el feedback para aprendizaje."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)
    data = json.loads(request.body)

    # Guardar feedback de IA para aprendizaje
    FeedbackIA.objects.create(
        ticket=ticket,
        corrector=request.user,
        ia_linea_original=ticket.linea_servicio or '',
        ia_proceso_original=ticket.proceso or '',
        ia_tipificacion_original=ticket.tipificacion or '',
        ia_criticidad_original=ticket.criticidad or '',
        linea_corregida=data.get('linea', ''),
        proceso_corregido=data.get('proceso', ''),
        tipificacion_corregida=data.get('tipificacion', ''),
        criticidad_corregida=data.get('criticidad', ''),
        observacion=data.get('observacion', '')
    )

    # Aplicar nueva clasificación
    if data.get('linea'): ticket.linea_servicio = data['linea']
    if data.get('proceso'): ticket.proceso = data['proceso']
    if data.get('tipificacion'): ticket.tipificacion = data['tipificacion']
    if data.get('criticidad'): ticket.criticidad = data['criticidad']
    if data.get('analisis_ia'): ticket.analisis_ia = data['analisis_ia']
    ticket.clasificado_por_ia = True
    ticket.save()

    LogActividad.objects.create(
        ticket=ticket, usuario=request.user,
        accion='Reclasificación IA aplicada y feedback registrado',
        detalle=f'Nueva tipificación: {data.get("tipificacion")} | Criticidad: {data.get("criticidad")}'
    )

    return JsonResponse({'ok': True})


# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN SLA
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def configurar_sla_view(request):
    """Panel para parametrizar el SLA. Los destinatarios se obtienen de los usuarios internos."""
    config = ConfiguracionSLA.objects.filter(activo=True).first()
    mensaje = None

    if request.method == 'POST':
        dias_peligro = int(request.POST.get('dias_alerta_peligro', 11))
        dias_vencido = int(request.POST.get('dias_alerta_vencido', 15))

        if config:
            config.dias_alerta_peligro = dias_peligro
            config.dias_alerta_vencido = dias_vencido
            config.save()
        else:
            config = ConfiguracionSLA.objects.create(
                dias_alerta_peligro=dias_peligro,
                dias_alerta_vencido=dias_vencido,
                activo=True
            )
        mensaje = {'tipo': 'exito', 'texto': 'Configuración SLA guardada exitosamente.'}

    # Destinatarios automáticos desde Usuarios Internos
    directores_regionales = PerfilUsuario.objects.filter(
        rol='director_regional'
    ).select_related('user').order_by('regional')

    admins_pqrs = PerfilUsuario.objects.filter(
        rol__in=['admin_pqrs', 'superadmin']
    ).select_related('user').order_by('user__first_name')

    return render(request, 'tickets/configurar_sla.html', {
        'config': config,
        'mensaje': mensaje,
        'perfil': request.user.perfil,
        'nav_active': 'sla',
        'directores_regionales': directores_regionales,
        'admins_pqrs': admins_pqrs,
    })


# ─────────────────────────────────────────────────────────────
# REPORTES Y ANALYTICS (ESTILO POWER BI)
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional')
def reportes_view(request):
    """Dashboard de analytics avanzado con KPIs interactivos."""
    perfil = request.user.perfil

    # Base query según rol
    if perfil.rol in ['superadmin', 'admin_pqrs']:
        qs = Ticket.objects.all()
    else:
        qs = Ticket.objects.filter(regional=perfil.regional)

    # ─ Filtros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    if fecha_inicio:
        qs = qs.filter(fecha_ingreso__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_ingreso__date__lte=fecha_fin)

    total = qs.count()
    resueltos = qs.filter(estado='resuelto').count()
    abiertos = qs.filter(estado='abierto').count()
    en_revision = qs.filter(estado='revision').count()
    cancelados = qs.filter(estado='cancelado').count()
    cumplimiento_sla = round((resueltos / total * 100), 1) if total > 0 else 0

    # Calcular vencidos sobre abiertos reales
    tickets_abiertos_list = list(qs.exclude(estado__in=['resuelto', 'cancelado']))
    vencidos_count = sum(1 for t in tickets_abiertos_list if t.estado_sla() == 'vencido')
    peligro_count = sum(1 for t in tickets_abiertos_list if t.estado_sla() == 'peligro')

    # ─ Por tipificación (top 8)
    por_tipificacion = list(
        qs.exclude(tipificacion__isnull=True).values('tipificacion')
        .annotate(total=Count('id')).order_by('-total')[:8]
    )

    # ─ Por regional
    por_regional = list(
        qs.exclude(regional__isnull=True).values('regional')
        .annotate(total=Count('id')).order_by('-total')
    )
    # Mapear códigos a labels
    regional_map = dict(Ticket.REGIONAL_CHOICES)
    for r in por_regional:
        r['label'] = regional_map.get(r['regional'], r['regional'])

    # ─ Por tipo de solicitud
    por_tipo = list(
        qs.values('tipo_solicitud').annotate(total=Count('id')).order_by('-total')
    )
    tipo_map = dict(Ticket.TIPO_SOLICITUD_CHOICES)
    for t in por_tipo:
        t['label'] = tipo_map.get(t['tipo_solicitud'], t['tipo_solicitud'])

    # ─ Por criticidad
    por_criticidad = list(
        qs.exclude(criticidad__isnull=True).values('criticidad')
        .annotate(total=Count('id')).order_by('-total')
    )
    criticidad_map = dict(Ticket.CRITICIDAD_CHOICES)
    for c in por_criticidad:
        c['label'] = criticidad_map.get(c['criticidad'], c['criticidad'])

    # ─ Tendencia mensual (últimos 6 meses)
    from django.db.models.functions import TruncMonth
    tendencia = list(
        qs.annotate(mes=TruncMonth('fecha_ingreso'))
        .values('mes').annotate(total=Count('id')).order_by('mes')
    )

    # ─ CSAT promedio
    encuestas = EncuestaSatisfaccion.objects.filter(
        ticket__in=qs, estado='satisfecho', puntuacion__isnull=False
    )
    csat_promedio = encuestas.aggregate(avg=Avg('puntuacion'))['avg']
    csat_promedio = round(csat_promedio, 1) if csat_promedio else None
    total_encuestas = encuestas.count()

    context = {
        'perfil': perfil,
        'total': total,
        'resueltos': resueltos,
        'abiertos': abiertos,
        'en_revision': en_revision,
        'cancelados': cancelados,
        'cumplimiento_sla': cumplimiento_sla,
        'vencidos_count': vencidos_count,
        'peligro_count': peligro_count,
        'csat_promedio': csat_promedio,
        'total_encuestas': total_encuestas,
        # Datos JSON para Chart.js
        'por_tipificacion_json': json.dumps(por_tipificacion, default=str),
        'por_regional_json': json.dumps(por_regional, default=str),
        'por_tipo_json': json.dumps(por_tipo, default=str),
        'por_criticidad_json': json.dumps(por_criticidad, default=str),
        'tendencia_json': json.dumps(tendencia, default=str),
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'nav_active': 'reportes',
    }
    return render(request, 'tickets/reportes.html', context)


# ─────────────────────────────────────────────────────────────
# EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs', 'director_regional')
def exportar_excel_view(request):
    """Exporta los tickets a Excel con formato corporativo Unidossis."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl no está instalado. Ejecute: pip install openpyxl', status=500)

    perfil = request.user.perfil
    if perfil.rol in ['superadmin', 'admin_pqrs']:
        qs = Ticket.objects.all()
    else:
        qs = Ticket.objects.filter(regional=perfil.regional)

    # Filtros opcionales
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    regional_filtro = request.GET.get('regional')
    if fecha_inicio:
        qs = qs.filter(fecha_ingreso__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_ingreso__date__lte=fecha_fin)
    if regional_filtro:
        qs = qs.filter(regional=regional_filtro)

    qs = qs.order_by('-fecha_ingreso')

    # ─ Crear libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte PQRS UNIDOSSIS'

    # ─ Colores corporativos
    AZUL_CORP = 'FF0F3460'
    ROJO_CORP = 'FFe63946'
    GRIS_HEADER = 'FF1E293B'
    GRIS_CLARO = 'FFF8FAFC'
    BLANCO = 'FFFFFFFF'

    # ─ Fila 1: Título corporativo
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = 'REPORTE DE PQRS — UNIDOSSIS S.A.S.'
    title_cell.font = Font(name='Calibri', size=14, bold=True, color=BLANCO)
    title_cell.fill = PatternFill('solid', fgColor=AZUL_CORP)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ─ Fila 2: Fecha del reporte
    ws.merge_cells('A2:O2')
    date_cell = ws['A2']
    date_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} | Por: {request.user.get_full_name() or request.user.username}'
    date_cell.font = Font(name='Calibri', size=10, color='FF64748B')
    date_cell.fill = PatternFill('solid', fgColor=GRIS_CLARO)
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # ─ Fila 3: Encabezados
    headers = [
        ('ID Caso', 12), ('Fecha Ingreso', 16), ('Tipo Solicitud', 16),
        ('Estado', 13), ('Criticidad', 14), ('Cliente / Institución', 28),
        ('Ciudad', 16), ('Solicitante', 22), ('Correo', 28),
        ('Regional', 18), ('Área / Proceso', 18), ('Línea de Servicio', 18),
        ('Tipificación', 24), ('Responsable', 20), ('Días Transcurridos', 10),
    ]

    thin = Side(style='thin', color='FFE2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header_text
        cell.font = Font(name='Calibri', size=10, bold=True, color=BLANCO)
        cell.fill = PatternFill('solid', fgColor=GRIS_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[3].height = 35

    # ─ Mapas de labels
    tipo_map = dict(Ticket.TIPO_SOLICITUD_CHOICES)
    estado_map = dict(Ticket.STATUS_CHOICES)
    criticidad_map = dict(Ticket.CRITICIDAD_CHOICES)
    regional_map = dict(Ticket.REGIONAL_CHOICES)
    area_map = dict(Ticket.AREA_CHOICES)
    linea_map = dict(Ticket.LINEA_CHOICES)
    tipificacion_map = dict(Ticket.TIPIFICACION_CHOICES)

    # ─ Colores de estado
    COLOR_ESTADO = {
        'abierto': 'FF3B82F6',
        'revision': 'FFF59E0B',
        'resuelto': 'FF10B981',
        'cancelado': 'FF94A3B8',
    }
    COLOR_CRITICIDAD = {
        'critica': 'FFEF4444',
        'mayor': 'FFF97316',
        'menor': 'FFF59E0B',
        'informativa': 'FF10B981',
    }

    # ─ Datos
    for row_idx, ticket in enumerate(qs, start=4):
        fill_row = PatternFill('solid', fgColor=BLANCO if row_idx % 2 == 0 else 'FFF8FAFC')
        row_data = [
            ticket.ticket_id,
            ticket.fecha_ingreso.strftime('%d/%m/%Y %H:%M'),
            tipo_map.get(ticket.tipo_solicitud, ticket.tipo_solicitud or ''),
            estado_map.get(ticket.estado, ticket.estado or ''),
            criticidad_map.get(ticket.criticidad, ticket.criticidad or ''),
            ticket.entidad_cliente or ticket.institucion or '',
            ticket.ciudad or '',
            ticket.remitente_nombre or '',
            ticket.remitente_email or '',
            regional_map.get(ticket.regional, ticket.regional or ''),
            area_map.get(ticket.proceso, ticket.proceso or ''),
            linea_map.get(ticket.linea_servicio, ticket.linea_servicio or ''),
            tipificacion_map.get(ticket.tipificacion, ticket.tipificacion or ''),
            ticket.responsable or '',
            ticket.dias_transcurridos(),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=9)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.fill = fill_row

            # Color especial para estado
            if col_idx == 4 and ticket.estado in COLOR_ESTADO:
                cell.font = Font(name='Calibri', size=9, bold=True, color=BLANCO)
                cell.fill = PatternFill('solid', fgColor=COLOR_ESTADO[ticket.estado])
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Color especial para criticidad
            if col_idx == 5 and ticket.criticidad in COLOR_CRITICIDAD:
                cell.font = Font(name='Calibri', size=9, bold=True, color=BLANCO)
                cell.fill = PatternFill('solid', fgColor=COLOR_CRITICIDAD[ticket.criticidad])
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A4'

    # ─ Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'PQRS_UNIDOSSIS_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────
# ENCUESTA CSAT
# ─────────────────────────────────────────────────────────────

def encuesta_csat_view(request, token):
    """Vista pública para que el cliente responda la encuesta de satisfacción."""
    encuesta = get_object_or_404(EncuestaSatisfaccion, token=token)

    if encuesta.estado in ['satisfecho', 'insatisfecho']:
        return render(request, 'tickets/encuesta_ya_respondida.html', {'encuesta': encuesta})

    if request.method == 'POST':
        puntuacion = int(request.POST.get('puntuacion', 0))
        comentario = request.POST.get('comentario', '').strip()

        encuesta.puntuacion = puntuacion
        encuesta.comentario_cliente = comentario
        encuesta.fecha_respuesta = timezone.now()

        ticket = encuesta.ticket

        if puntuacion >= 4:
            # Satisfecho → cierre definitivo
            encuesta.estado = 'satisfecho'
            ticket.estado = 'resuelto'
            ticket.save()
            LogActividad.objects.create(
                ticket=ticket, usuario=None,
                accion=f'Encuesta CSAT respondida: SATISFECHO ({puntuacion}/5 ⭐)',
                detalle=f'Comentario: {comentario[:100]}' if comentario else 'Sin comentario'
            )
        else:
            # Insatisfecho → reabrir y alertar
            encuesta.estado = 'insatisfecho'
            ticket.estado = 'revision'
            ticket.save()
            LogActividad.objects.create(
                ticket=ticket, usuario=None,
                accion=f'🚨 Encuesta CSAT: INSATISFECHO ({puntuacion}/5 ⭐) — Ticket REABIERTO automáticamente',
                detalle=f'Comentario cliente: {comentario[:200]}' if comentario else 'Sin comentario'
            )
            # Notificar al equipo que el cliente está insatisfecho
            try:
                config = ConfiguracionSLA.objects.filter(activo=True).first()
                destinatarios = config.get_emails_vencido() if config else []
                if destinatarios:
                    send_mail(
                        subject=f'🚨 CSAT NEGATIVO — Ticket {ticket.ticket_id} reabierto automáticamente',
                        message=(
                            f'El cliente {ticket.remitente_nombre or ticket.remitente_email} '
                            f'respondió la encuesta CSAT con una puntuación de {puntuacion}/5.\n\n'
                            f'El ticket #{ticket.ticket_id} ha sido REABIERTO automáticamente y '
                            f'requiere revisión urgente.\n\n'
                            f'Comentario del cliente: {comentario or "(Sin comentario)"}\n\n'
                            f'— UNIDOSSIS PQRS · Sistema Automático'
                        ),
                        from_email='UNIDOSSIS PQRS <alertas@unidossis.com.co>',
                        recipient_list=destinatarios,
                        fail_silently=True,
                    )
            except Exception:
                pass

        encuesta.save()
        return render(request, 'tickets/encuesta_gracias.html', {
            'encuesta': encuesta,
            'satisfecho': encuesta.estado == 'satisfecho'
        })

    return render(request, 'tickets/encuesta_csat.html', {'encuesta': encuesta})


# ─────────────────────────────────────────────────────────────
# MONITOREO DEL SISTEMA
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin')
def monitoreo_view(request):
    """Panel de monitoreo: muestra logs de rendimiento y errores del sistema."""
    import os
    from pathlib import Path

    log_dir = Path(__file__).resolve().parent.parent / 'logs'

    def leer_ultimas_lineas(archivo, n=100):
        """Lee las últimas N líneas de un archivo de log."""
        ruta = log_dir / archivo
        if not ruta.exists():
            return []
        try:
            with open(ruta, 'r', encoding='utf-8', errors='replace') as f:
                lineas = f.readlines()
            return [l.strip() for l in lineas[-n:] if l.strip()][::-1]  # Más reciente primero
        except Exception:
            return []

    log_rendimiento = leer_ultimas_lineas('rendimiento.log', 150)
    log_errores = leer_ultimas_lineas('errores.log', 100)

    # Estadísticas
    total_peticiones = len(log_rendimiento)
    peticiones_lentas = sum(1 for l in log_rendimiento if 'LENTO' in l)
    total_errores = len(log_errores)

    # Tamaño de la base de datos
    db_path = Path(__file__).resolve().parent.parent / 'db.sqlite3'
    if db_path.exists():
        tamano_bytes = db_path.stat().st_size
        if tamano_bytes > 1024 * 1024:
            tamano_db = f'{tamano_bytes / (1024*1024):.1f} MB'
        else:
            tamano_db = f'{tamano_bytes / 1024:.0f} KB'
    else:
        tamano_db = 'N/A'

    return render(request, 'tickets/monitoreo.html', {
        'log_rendimiento': log_rendimiento,
        'log_errores': log_errores,
        'total_peticiones': total_peticiones,
        'peticiones_lentas': peticiones_lentas,
        'total_errores': total_errores,
        'tamano_db': tamano_db,
        'perfil': request.user.perfil,
        'nav_active': 'monitoreo',
    })


@login_required
@rol_requerido('superadmin')
def descargar_log_view(request, tipo):
    """Descarga un archivo de log específico."""
    from pathlib import Path

    log_dir = Path(__file__).resolve().parent.parent / 'logs'
    archivos_permitidos = {
        'rendimiento': 'rendimiento.log',
        'errores': 'errores.log',
    }

    if tipo not in archivos_permitidos:
        return HttpResponse('Archivo no válido.', status=400)

    ruta = log_dir / archivos_permitidos[tipo]
    if not ruta.exists():
        return HttpResponse(f'El archivo {archivos_permitidos[tipo]} aún no tiene registros.', status=404)

    with open(ruta, 'r', encoding='utf-8', errors='replace') as f:
        contenido = f.read()

    fecha = timezone.now().strftime('%Y-%m-%d_%H%M')
    response = HttpResponse(contenido, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="unidossis_{tipo}_{fecha}.log"'
    return response


@login_required
@rol_requerido('superadmin')
def descargar_respaldo_db_view(request):
    """Descarga un respaldo completo de la base de datos SQLite."""
    from pathlib import Path

    db_path = Path(__file__).resolve().parent.parent / 'db.sqlite3'
    if not db_path.exists():
        return HttpResponse('Base de datos no encontrada.', status=404)

    fecha = timezone.now().strftime('%Y-%m-%d_%H%M')
    with open(db_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/x-sqlite3')
        response['Content-Disposition'] = f'attachment; filename="unidossis_respaldo_{fecha}.sqlite3"'
    return response


# ─────────────────────────────────────────────────────────────
# CREAR PQRS MANUAL (desde el dashboard)
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin', 'admin_pqrs')
def crear_pqrs_manual_view(request):
    """Permite a superadmin y admin_pqrs crear una PQRS manualmente
    para casos recibidos por teléfono, presencial, etc."""
    if request.method != 'POST':
        return redirect('dashboard')

    institucion = request.POST.get('institucion', '')
    ciudad = request.POST.get('ciudad', '')
    nombre = request.POST.get('nombre', '')
    cargo = request.POST.get('cargo', '')
    telefono = request.POST.get('telefono', '')
    email = request.POST.get('email', '')
    tipo = request.POST.get('tipo_solicitud', 'queja')
    regional = request.POST.get('regional', '')
    asunto = request.POST.get('asunto', '')
    descripcion = request.POST.get('descripcion', '')
    medio_recepcion = request.POST.get('medio_recepcion', 'manual')

    if not (institucion and nombre and email and asunto and descripcion):
        from django.contrib import messages
        messages.error(request, 'Complete todos los campos obligatorios.')
        return redirect('dashboard')

    # Buscar cliente existente
    cliente_rel = Cliente.objects.filter(nombre__iexact=institucion).first()

    # Procesamiento por IA
    analisis = analizar_ticket_con_ia(asunto, descripcion)
    regional_asignada = regional if regional else (cliente_rel.regional if cliente_rel else analisis.get('regional', 'liquidos'))

    # Autocrear cliente si no existe
    if not cliente_rel:
        ciudad_obj = Ciudad.objects.filter(nombre__iexact=ciudad).first() if ciudad else None
        cliente_rel = Cliente.objects.create(
            nombre=institucion,
            regional=regional_asignada,
            ciudad=ciudad_obj,
            email_principal=email,
            activo=True
        )

    nuevo_ticket = Ticket.objects.create(
        cliente_rel=cliente_rel,
        entidad_cliente=institucion,
        institucion=institucion,
        ciudad=ciudad,
        remitente_nombre=nombre,
        solicitante_cargo=cargo,
        telefono=telefono,
        remitente_email=email,
        tipo_solicitud=tipo,
        asunto=asunto,
        cuerpo=descripcion,
        regional=regional_asignada,
        proceso=analisis['proceso'],
        linea_servicio=analisis['linea'],
        tipificacion=analisis['tipificacion'],
        criticidad=analisis['criticidad'],
        analisis_ia=analisis['analisis_ia'],
        clasificado_por_ia=True
    )

    # Archivos adjuntos
    if request.FILES.getlist('adjuntos'):
        for f in request.FILES.getlist('adjuntos'):
            ArchivoAdjunto.objects.create(ticket=nuevo_ticket, archivo=f)

    # Log de actividad
    LogActividad.objects.create(
        ticket=nuevo_ticket, usuario=request.user,
        accion=f'PQRS creada manualmente ({medio_recepcion})',
        detalle=f'Creado por {request.user.get_full_name() or request.user.username} | '
                f'Medio: {medio_recepcion} | IA: {analisis["tipificacion"]} / {analisis["criticidad"]}'
    )

    return redirect('ticket_detail', ticket_id=nuevo_ticket.ticket_id)


# ─────────────────────────────────────────────────────────────
# API BUSCAR TICKETS - BUSCADOR INTELIGENTE DASHBOARD
# ─────────────────────────────────────────────────────────────

@login_required
def api_buscar_tickets(request):
    """Busca coincidencias agrupadas por categoría para el buscador inteligente."""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'grupos': []})

    perfil = request.user.perfil
    # Base query según rol
    if perfil.rol in ['superadmin', 'admin_pqrs']:
        qs = Ticket.objects.all()
    elif perfil.rol == 'director_regional':
        qs = Ticket.objects.filter(regional=perfil.regional)
    elif perfil.rol == 'supervisor':
        filtro = Q(responsable__icontains=request.user.username)
        if request.user.get_full_name():
            filtro |= Q(responsable__icontains=request.user.get_full_name())
        if request.user.last_name:
            filtro |= Q(responsable__icontains=request.user.last_name)
        qs = Ticket.objects.filter(filtro)
    else:
        return JsonResponse({'grupos': []})

    grupos = []

    # ── 1. CLIENTES (primera prioridad)
    # Solo clientes que tienen tickets visibles para este usuario (respeta rol)
    clientes_con_tickets_ids = qs.exclude(
        cliente_rel__isnull=True
    ).values_list('cliente_rel_id', flat=True).distinct()

    clientes_match = Cliente.objects.filter(
        id__in=clientes_con_tickets_ids,
        nombre__icontains=q
    ).distinct()[:6]

    items_clientes = []
    seen_clientes = set()
    for c in clientes_match:
        cant = qs.filter(cliente_rel=c).count()
        if cant > 0 and c.id not in seen_clientes:
            seen_clientes.add(c.id)
            items_clientes.append({
                'tipo': 'cliente',
                'texto': c.nombre,
                'subtexto': c.get_regional_display() if c.regional else 'Sin regional',
                'cantidad': cant,
                'accion': f'/dashboard/?cliente_id={c.id}',
                'id': c.id,
            })

    # También buscar entidades libres (tickets sin cliente_rel) que coincidan
    entidades = qs.filter(
        cliente_rel__isnull=True
    ).filter(
        Q(entidad_cliente__icontains=q) | Q(institucion__icontains=q)
    ).values_list('entidad_cliente', flat=True).distinct()[:4]

    seen_ent = set(c['texto'].lower() for c in items_clientes)
    for e in entidades:
        if e and e.strip() and e.lower() not in seen_ent:
            seen_ent.add(e.lower())
            cant = qs.filter(Q(entidad_cliente__iexact=e) | Q(institucion__iexact=e)).count()
            items_clientes.append({
                'tipo': 'entidad',
                'texto': e,
                'subtexto': 'Institución en tickets',
                'cantidad': cant,
                'accion': f'/dashboard/?q={e}',
            })

    if items_clientes:
        grupos.append({
            'titulo': 'Clientes / Instituciones',
            'icono': 'fa-hospital',
            'color': '#059669',
            'items': items_clientes,
        })

    # ── 2. RESPONSABLES
    responsables = qs.filter(responsable__icontains=q).exclude(
        responsable__isnull=True
    ).exclude(responsable='').values_list('responsable', flat=True).distinct()[:4]
    if responsables:
        items = []
        seen_r = set()
        for r in responsables:
            r_clean = r.strip()
            if r_clean.lower() not in seen_r:
                seen_r.add(r_clean.lower())
                cant = qs.filter(responsable__iexact=r_clean).count()
                items.append({
                    'tipo': 'responsable',
                    'texto': r_clean,
                    'subtexto': 'Personal asignado',
                    'cantidad': cant,
                    'accion': f'/dashboard/?q={r_clean}',
                })
        if items:
            grupos.append({
                'titulo': 'Responsables',
                'icono': 'fa-user-tie',
                'color': '#dc2626',
                'items': items,
            })

    # ── 3. REGIONALES
    for key, nombre in Ticket.REGIONAL_CHOICES:
        if q.lower() in nombre.lower() or q.lower() in key.lower():
            cant = qs.filter(regional=key).count()
            if cant > 0:
                if not any(g['titulo'] == 'Regionales' for g in grupos):
                    grupos.append({
                        'titulo': 'Regionales',
                        'icono': 'fa-map-location-dot',
                        'color': '#0ea5e9',
                        'items': [],
                    })
                for g in grupos:
                    if g['titulo'] == 'Regionales':
                        g['items'].append({
                            'tipo': 'regional',
                            'texto': nombre,
                            'subtexto': f'Código: {key}',
                            'cantidad': cant,
                            'accion': f'/dashboard/?q={nombre}',
                        })

    # ── 4. TICKETS ESPECÍFICOS (por ID)
    if 'PQRS' in q.upper() or q.upper().startswith('#'):
        tickets_match = qs.filter(ticket_id__icontains=q.replace('#', ''))[:5]
        if tickets_match:
            items = []
            for t in tickets_match:
                items.append({
                    'tipo': 'ticket',
                    'texto': t.ticket_id,
                    'subtexto': f'{t.remitente_nombre or t.entidad_cliente or "—"} · {t.asunto[:40] if t.asunto else "—"}',
                    'cantidad': None,
                    'accion': f'/ticket/{t.ticket_id}/',
                    'estado': t.get_estado_display(),
                    'sla': t.estado_sla(),
                })
            grupos.append({
                'titulo': 'Tickets',
                'icono': 'fa-hashtag',
                'color': '#3b82f6',
                'items': items,
            })

    # ── 5. COINCIDENCIAS POR ASUNTO (últimos)
    tickets_asunto = qs.filter(asunto__icontains=q)[:4]
    if tickets_asunto:
        items = []
        for t in tickets_asunto:
            items.append({
                'tipo': 'ticket',
                'texto': t.asunto[:55] if t.asunto else '—',
                'subtexto': f'{t.ticket_id} · {t.remitente_nombre or t.entidad_cliente or "—"}',
                'cantidad': None,
                'accion': f'/ticket/{t.ticket_id}/',
                'estado': t.get_estado_display(),
                'sla': t.estado_sla(),
            })
        grupos.append({
            'titulo': 'Coincidencias en Asunto',
            'icono': 'fa-align-left',
            'color': '#d97706',
            'items': items,
        })

    return JsonResponse({'grupos': grupos})


# ─────────────────────────────────────────────────────────────
# CONTROL DE CAMBIOS (Changelog / Version Control)
# ─────────────────────────────────────────────────────────────

@login_required
@rol_requerido('superadmin')
def control_cambios_view(request):
    """Panel de control de cambios: muestra historial de commits, estado de sincronización
    con producción y permite revertir cambios específicos."""
    import subprocess
    from pathlib import Path

    repo_dir = Path(__file__).resolve().parent.parent.parent  # raíz del repo

    def run_git(args, cwd=None):
        """Ejecuta un comando git y retorna la salida."""
        try:
            result = subprocess.run(
                ['git'] + args,
                cwd=str(cwd or repo_dir),
                capture_output=True, text=True, timeout=15,
                encoding='utf-8', errors='replace'
            )
            return result.stdout.strip()
        except Exception:
            return ''

    # Obtener commits recientes (últimos 30)
    log_output = run_git([
        'log', '--pretty=format:%H||%h||%an||%ad||%s', '--date=format:%d/%m/%Y %H:%M', '-30'
    ])

    commits = []
    for line in (log_output.split('\n') if log_output else []):
        parts = line.split('||')
        if len(parts) >= 5:
            msg = parts[4]
            # Detectar tipo de commit por prefijo convencional (Conventional Commits spec)
            prefix_type = 'other'
            msg_lower = msg.lower().strip()
            if msg_lower.startswith('feat') or '✨' in msg:
                prefix_type = 'feat'
            elif msg_lower.startswith('fix') or '🐛' in msg or 'fix:' in msg_lower:
                prefix_type = 'fix'
            elif msg_lower.startswith('docs') or '📝' in msg:
                prefix_type = 'docs'
            elif msg_lower.startswith('style') or '💄' in msg:
                prefix_type = 'style'
            elif msg_lower.startswith('refactor') or '♻️' in msg:
                prefix_type = 'refactor'
            elif msg_lower.startswith('perf') or '⚡' in msg:
                prefix_type = 'perf'
            elif msg_lower.startswith('test') or '✅' in msg:
                prefix_type = 'test'
            elif msg_lower.startswith(('ci', 'build', 'config')) or '👷' in msg:
                prefix_type = 'ci'
            elif msg_lower.startswith('chore') or '🔧' in msg:
                prefix_type = 'chore'
            elif msg_lower.startswith('revert'):
                prefix_type = 'revert'

            commits.append({
                'hash': parts[0],
                'hash_short': parts[1],
                'author': parts[2],
                'date': parts[3],
                'message': msg,
                'prefix_type': prefix_type,
                'is_in_prod': False,  # Se actualiza abajo
            })

    # Comparar con origin/main para detectar qué está en producción
    commits_ahead = 0
    try:
        # Fetch silencioso para actualizar la referencia remota
        run_git(['fetch', '--quiet'])
        ahead_output = run_git(['rev-list', '--count', 'origin/main..HEAD'])
        commits_ahead = int(ahead_output) if ahead_output.isdigit() else 0

        # Obtener el hash del último commit en producción
        prod_hash = run_git(['rev-parse', 'origin/main'])

        # Marcar commits que ya están en producción
        for i, commit in enumerate(commits):
            if i >= commits_ahead:
                commit['is_in_prod'] = True
    except Exception:
        pass

    # Estado de sincronización
    sync_status = 'synced' if commits_ahead == 0 else 'ahead'

    # Cambios pendientes (no commiteados)
    status_output = run_git(['status', '--short'])
    cambios_pendientes = [l.strip() for l in status_output.split('\n') if l.strip()] if status_output else []

    # Archivos modificados localmente
    archivos_modificados = len(cambios_pendientes)

    # Fecha del último commit
    ultimo_commit_fecha = commits[0]['date'].split(' ')[0] if commits else 'N/A'

    # ── Versión actual del proyecto ──────────────────────────
    version_actual = 'v1.0.0'
    try:
        from tickets import __version__
        version_actual = f'v{__version__}'
    except Exception:
        pass

    # ── Leer CHANGELOG.md (primeras secciones) ──────────────
    changelog_preview = ''
    changelog_path = repo_dir / 'CHANGELOG.md'
    if changelog_path.exists():
        try:
            with open(changelog_path, 'r', encoding='utf-8') as f:
                changelog_preview = f.read(3000)  # Primeras 3000 chars
        except Exception:
            changelog_preview = ''

    # ── Releases/Tags del repositorio ───────────────────────
    tags_output = run_git(['tag', '-l', '--sort=-version:refname', '--format=%(refname:short)|||%(creatordate:short)|||%(subject)'])
    releases = []
    for line in (tags_output.split('\n') if tags_output else []):
        parts = line.split('|||')
        if parts and parts[0].strip():
            releases.append({
                'tag': parts[0].strip(),
                'fecha': parts[1].strip() if len(parts) > 1 else '',
                'descripcion': parts[2].strip() if len(parts) > 2 else '',
            })

    # ── Estado del último CI run (desde GitHub Actions API) ─
    # Leemos el estado via requests a la API pública del repo
    ci_status = 'unknown'
    ci_url = 'https://github.com/FRANCISCOJOSEALVAREZABUABARA/unidossis-pqrs/actions'
    try:
        import urllib.request
        api_url = 'https://api.github.com/repos/FRANCISCOJOSEALVAREZABUABARA/unidossis-pqrs/actions/runs?branch=main&per_page=1'
        req = urllib.request.Request(api_url, headers={'User-Agent': 'UnidossisPQRS/1.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            import json as _json
            data = _json.loads(resp.read())
            if data.get('workflow_runs'):
                run = data['workflow_runs'][0]
                ci_status = run.get('conclusion') or run.get('status', 'unknown')
                ci_url = run.get('html_url', ci_url)
    except Exception:
        pass  # No bloquear si GitHub no responde

    context = {
        'commits': commits,
        'total_commits': len(commits),
        'commits_ahead': commits_ahead,
        'sync_status': sync_status,
        'cambios_pendientes': cambios_pendientes,
        'archivos_modificados': archivos_modificados,
        'ultimo_commit_fecha': ultimo_commit_fecha,
        'version_actual': version_actual,
        'changelog_preview': changelog_preview,
        'releases': releases,
        'ci_status': ci_status,
        'ci_url': ci_url,
        'perfil': request.user.perfil,
        'nav_active': 'control_cambios',
    }
    return render(request, 'tickets/control_cambios.html', context)


@login_required
@rol_requerido('superadmin')
def api_detalle_commit(request, commit_hash):
    """Retorna los archivos modificados en un commit específico (vía AJAX)."""
    import subprocess
    from pathlib import Path

    repo_dir = Path(__file__).resolve().parent.parent.parent

    # Validar hash (seguridad)
    if not all(c in '0123456789abcdefABCDEF' for c in commit_hash):
        return JsonResponse({'error': 'Hash inválido.'}, status=400)

    try:
        result = subprocess.run(
            ['git', 'diff-tree', '--no-commit-id', '-r', '--name-status', commit_hash],
            cwd=str(repo_dir), capture_output=True, text=True, timeout=10,
            encoding='utf-8', errors='replace'
        )
        files = []
        for line in result.stdout.strip().split('\n'):
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 2:
                    files.append({
                        'status': parts[0][0],  # M, A, D, R
                        'path': parts[-1],
                    })
        return JsonResponse({'files': files})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@rol_requerido('superadmin')
def api_diff_commit(request, commit_hash):
    """Retorna el diff de un commit específico (vía AJAX)."""
    import subprocess
    from pathlib import Path

    repo_dir = Path(__file__).resolve().parent.parent.parent

    # Validar hash
    if not all(c in '0123456789abcdefABCDEF' for c in commit_hash):
        return JsonResponse({'error': 'Hash inválido.'}, status=400)

    try:
        result = subprocess.run(
            ['git', 'diff', f'{commit_hash}~1', commit_hash, '--stat=120', '--no-color'],
            cwd=str(repo_dir), capture_output=True, text=True, timeout=10,
            encoding='utf-8', errors='replace'
        )
        # También obtener diff con contexto limitado
        diff_result = subprocess.run(
            ['git', 'diff', f'{commit_hash}~1', commit_hash, '--no-color', '-U3'],
            cwd=str(repo_dir), capture_output=True, text=True, timeout=10,
            encoding='utf-8', errors='replace'
        )
        diff_lines = diff_result.stdout.split('\n')[:500]  # Limitar a 500 líneas
        return JsonResponse({'diff': diff_lines, 'stat': result.stdout})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@rol_requerido('superadmin')
def api_revertir_commit(request, commit_hash):
    """Revierte un commit específico creando un nuevo commit de reversión.
    Solo funciona para commits que no están en producción."""
    import subprocess
    from pathlib import Path

    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)

    repo_dir = Path(__file__).resolve().parent.parent.parent

    # Validar hash
    if not all(c in '0123456789abcdefABCDEF' for c in commit_hash):
        return JsonResponse({'error': 'Hash inválido.'}, status=400)

    try:
        # Verificar que el commit no está en producción
        ahead_output = subprocess.run(
            ['git', 'rev-list', '--count', 'origin/main..HEAD'],
            cwd=str(repo_dir), capture_output=True, text=True, timeout=10,
            encoding='utf-8', errors='replace'
        )
        commits_ahead = int(ahead_output.stdout.strip()) if ahead_output.stdout.strip().isdigit() else 0

        local_commits = subprocess.run(
            ['git', 'rev-list', 'origin/main..HEAD'],
            cwd=str(repo_dir), capture_output=True, text=True, timeout=10,
            encoding='utf-8', errors='replace'
        )
        local_hashes = local_commits.stdout.strip().split('\n') if local_commits.stdout.strip() else []

        if commit_hash not in local_hashes:
            return JsonResponse({
                'success': False,
                'error': 'Este commit ya está en producción y no se puede revertir desde aquí.'
            })

        # Ejecutar revert
        result = subprocess.run(
            ['git', 'revert', '--no-edit', commit_hash],
            cwd=str(repo_dir), capture_output=True, text=True, timeout=30,
            encoding='utf-8', errors='replace'
        )

        if result.returncode == 0:
            # Log de auditoría
            LogActividad.objects.create(
                ticket=None, usuario=request.user,
                accion=f'Commit revertido: {commit_hash[:7]}',
                detalle=f'Revert ejecutado por: {request.user.get_full_name() or request.user.username}'
            )
            return JsonResponse({
                'success': True,
                'message': f'Commit {commit_hash[:7]} revertido exitosamente. Se creó un nuevo commit de reversión.'
            })
        else:
            # Si hay conflictos, abortar
            subprocess.run(
                ['git', 'revert', '--abort'],
                cwd=str(repo_dir), capture_output=True, text=True, timeout=10
            )
            return JsonResponse({
                'success': False,
                'error': f'No se pudo revertir automáticamente (conflictos). Detalle: {result.stderr[:200]}'
            })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ─────────────────────────────────────────────────────────────
# VIEW AS / ROLE IMPERSONATION (Solo superadmin)
# ─────────────────────────────────────────────────────────────

@login_required
def api_simular_rol(request):
    """
    Activa/desactiva la simulación de rol para el superadmin.

    POST { "rol": "director_regional", "regional": "llanos" }
    POST { "rol": "cliente", "cliente_id": 42 }
    POST { "rol": "agente" }
    POST { "rol": "" }  → desactiva
    """
    # Verificar que el usuario REAL es superadmin (no simulado)
    perfil_real_rol = request.rol_original or request.user.perfil.rol
    rol_real_session = request.session.get('_rol_real_superadmin', False)

    if perfil_real_rol != 'superadmin' and not rol_real_session:
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        data = {}

    rol = data.get('rol', '').strip()
    roles_validos = ['admin_pqrs', 'director_regional', 'agente', 'cliente']

    if rol and rol in roles_validos:
        request.session['simular_rol'] = rol
        request.session['_rol_real_superadmin'] = True

        # ── Guardar regional si es director_regional ──
        if rol == 'director_regional':
            regional = data.get('regional', '').strip()
            request.session['simular_regional'] = regional

        # ── Guardar cliente_id si es cliente ──
        if rol == 'cliente':
            cliente_id = data.get('cliente_id')
            if cliente_id:
                request.session['simular_cliente_id'] = cliente_id
            else:
                request.session.pop('simular_cliente_id', None)

        # Determinar redirección
        if rol == 'cliente':
            redirect_url = '/cliente/dashboard/'
        else:
            redirect_url = '/dashboard/'

        rol_labels = {
            'admin_pqrs': 'Administrador PQRS',
            'director_regional': 'Director Regional',
            'agente': 'Agente / Consultor',
            'cliente': 'Cliente Institución',
        }

        return JsonResponse({
            'ok': True,
            'activo': True,
            'rol': rol,
            'rol_display': rol_labels.get(rol, rol),
            'redirect': redirect_url,
        })

    elif rol == '' or not rol:
        # Desactivar simulación — limpiar todo
        request.session.pop('simular_rol', None)
        request.session.pop('simular_regional', None)
        request.session.pop('simular_cliente_id', None)
        request.session.pop('_rol_real_superadmin', None)
        return JsonResponse({
            'ok': True,
            'activo': False,
            'redirect': '/dashboard/',
        })
    else:
        return JsonResponse({'ok': False, 'error': 'Rol no válido.'}, status=400)


@login_required
def api_simular_opciones(request):
    """
    Devuelve las opciones disponibles para sub-selects de simulación:
    - Regionales disponibles (para director_regional)
    - Clientes disponibles (para cliente)
    Solo accesible por superadmin.
    """
    if request.user.perfil.rol != 'superadmin' and not request.session.get('_rol_real_superadmin'):
        return JsonResponse({'ok': False}, status=403)

    regionales = Ticket.REGIONAL_CHOICES

    clientes = list(
        Cliente.objects.filter(activo=True).values('id', 'nombre', 'regional')
        .order_by('nombre')[:80]
    )

    return JsonResponse({
        'ok': True,
        'regionales': [{'key': k, 'label': v} for k, v in regionales],
        'clientes': clientes,
    })

